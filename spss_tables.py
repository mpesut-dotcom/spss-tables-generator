#!/usr/bin/env python3
"""
spss_tables.py — Python zamjena za SPSS 15 TABLES proceduru
============================================================

Cita isti input.txt format koji koriste SBS skripte i generira
tablice u Excel formatu iz .sav podatkovnog fajla.

Koristenje:
    python spss_tables.py --sav data.sav --input input.txt --output output.xlsx
    python spss_tables.py --sav data.sav --input input.txt --output output.xlsx --start 1
    python spss_tables.py --sav data.sav --input input.txt --weight ponder --filter filter_var

Instalacija potrebnih paketa:
    pip install pyreadstat pandas openpyxl

Podrzani tipovi tablica (definirani u input.txt):
    s  = Simple/Basic frequency tablica (n, %)
    k  = Multiple Response - MRGROUP (n, %, baza=ispitanici)
    d  = Multiple Dichotomy - MDGROUP (n, %, baza=ispitanici)
    n  = Numericke statistike (Mean, StdDev, Median, Min, Max, N)
    m  = Numericke statistike skraceno (Mean, N)
    f  = Frequencies sortirane po frekvenciji silazno
"""

import argparse
import os
import sys
import time

import numpy as np
import pandas as pd
import pyreadstat
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side,
)
from openpyxl.utils import get_column_letter


# ═══════════════════════════════════════════════════════════════════
#  PARSIRANJE INPUT FAJLA
# ═══════════════════════════════════════════════════════════════════

def parse_input_file(filepath):
    """
    Parsira input.txt s 3 sekcije razdvojene praznim linijama:
      Sekcija 1: Break varijable (ob[])
      Sekcija 2: Naslovi s type prefiksom (naslov[])
      Sekcija 3: Definicije varijabli (pitanje[])
    Vraca: (break_vars, titles, variables)
    """
    for enc in ('utf-8', 'cp1250', 'latin-1'):
        try:
            with open(filepath, 'r', encoding=enc) as f:
                content = f.read()
            break
        except UnicodeDecodeError:
            continue

    lines = content.replace('\r\n', '\n').split('\n')

    sections = []
    current = []
    for line in lines:
        if line.strip() == '':
            if current:
                sections.append(current)
                current = []
        else:
            current.append(line)
    if current:
        sections.append(current)

    if len(sections) < 3:
        raise ValueError(
            f"Input fajl mora imati 3 sekcije razdvojene blank linijama, "
            f"pronadjeno: {len(sections)}"
        )

    return sections[0], sections[1], sections[2]


# ═══════════════════════════════════════════════════════════════════
#  POMOCNE FUNKCIJE
# ═══════════════════════════════════════════════════════════════════

def build_column_map(df):
    """Case-insensitive mapa imena kolona (SPSS je case-insensitive)."""
    return {c.lower(): c for c in df.columns}


def resolve_col(name, df, col_map):
    """Razrijesi ime varijable case-insensitive."""
    name = name.strip()
    if name in df.columns:
        return name
    lower = name.lower()
    if lower in col_map:
        return col_map[lower]
    raise KeyError(name)


def get_var_label(var_name, meta):
    """Dohvati label varijable iz metapodataka."""
    labels = getattr(meta, 'column_names_to_labels', {}) or {}
    return labels.get(var_name) or var_name


def get_value_labels(var_name, meta):
    """Dohvati value labels za varijablu."""
    val_labels = getattr(meta, 'variable_value_labels', {}) or {}
    return val_labels.get(var_name, {})


def label_for_value(val, val_labels):
    """Pronadi label za vrijednost, handling float/int key mismatch."""
    if val in val_labels:
        return val_labels[val]
    try:
        fval = float(val)
        if fval in val_labels:
            return val_labels[fval]
    except (ValueError, TypeError):
        pass
    try:
        ival = int(float(val))
        if ival in val_labels:
            return val_labels[ival]
    except (ValueError, TypeError):
        pass
    # Lijepo formatiraj (makni .0 od floatova)
    if isinstance(val, float) and val == int(val):
        return str(int(val))
    return str(val)


def get_table_type(title_line):
    """Izvuci tip tablice iz prva 2 znaka naslova."""
    return title_line[:2].strip()


def get_table_title(title_line):
    """Izvuci naslov tablice (bez type prefiksa)."""
    return title_line[2:].strip()


def parse_mr_vars(var_string):
    """
    Parsira MR variable string.
    Format: $e1 '' var1 var2 var3 ...
    Vraca listu stvarnih imena varijabli (bez $e1 i '').
    """
    parts = var_string.strip().split()
    var_names = [p for p in parts if not p.startswith('$') and p != "''"]
    return var_names


def parse_numeric_vars(var_string):
    """
    Parsira variable string za 'n' i 'm' tipove.
    String se dijeli tocno na pola (po broju znakova), isto kao VB Left/Right.
    Lijeva strana: space-separated observation variable names
    Desna strana: iste varijable spojene s '+' (ignoriramo)
    """
    mid = len(var_string) // 2
    left_half = var_string[:mid].strip()
    return left_half.split()


# ═══════════════════════════════════════════════════════════════════
#  GENERIRANJE TABLICA
# ═══════════════════════════════════════════════════════════════════

def make_simple_table(df, var_name, meta, col_map, weight_col=None):
    """
    Tip 's': Basic frequency tablica s n i %.
    Replicira: TABLES /FORMAT ZERO MISSING('.') /FTOTAL $t 'Total'
               /TABLES (var + $t) BY (STATISTICS)
               /STATISTICS count((F5.0) 'n') cpct((F5.1) '%')
    """
    actual_col = resolve_col(var_name, df, col_map)
    subset = df[df[actual_col].notna()].copy()
    val_labels = get_value_labels(actual_col, meta)

    if weight_col and weight_col in subset.columns:
        counts = subset.groupby(actual_col)[weight_col].sum()
    else:
        counts = subset[actual_col].value_counts(sort=False)

    # Sakupi sve kategorije (iz podataka + value labels), sortiraj
    all_values = set(counts.index)
    if val_labels:
        all_values.update(val_labels.keys())

    def _sort_key(val):
        try:
            return (0, float(val))
        except (ValueError, TypeError):
            return (1, str(val))

    all_values = sorted(all_values, key=_sort_key)

    total_n = counts.sum()

    rows = []
    for val in all_values:
        lbl = label_for_value(val, val_labels) if val_labels else label_for_value(val, {})
        n = counts.get(val, 0)
        pct = round(float(n) / float(total_n) * 100, 1) if total_n > 0 else 0.0
        rows.append((str(lbl), round(float(n), 1), pct))

    rows.append(('Total', round(float(total_n), 1), 100.0))

    return {
        'header': ['', 'n', '%'],
        'rows': rows,
        'caption': '',
    }


def make_mr_table(df, var_string, meta, col_map, mr_type='k', weight_col=None):
    """
    Tip 'k'/'d': Multiple response tablica.
    'd' (MDGROUP): svaka varijabla = 1 red, label = variable label
    'k' (MRGROUP): redovi = unique value labels across all variables
    """
    var_names_raw = parse_mr_vars(var_string)
    var_names = [resolve_col(v, df, col_map) for v in var_names_raw]

    # Total cases: ispitanici s barem jednim ne-missing odgovorom
    mr_df = df[var_names]
    valid_mask = mr_df.notna().any(axis=1)
    use_weight = weight_col and weight_col in df.columns

    if use_weight:
        total_cases = float(df.loc[valid_mask, weight_col].sum())
    else:
        total_cases = int(valid_mask.sum())

    rows = []

    if mr_type == 'k':
        # MRGROUP: rows = unique values, labels = value labels
        # Collect all unique values across all k-variables
        all_vals = set()
        for vname in var_names:
            all_vals.update(df[vname].dropna().unique())
        all_vals.discard(0)
        all_vals = sorted(all_vals,
                          key=lambda x: (0, float(x)) if isinstance(x, (int, float)) else (1, str(x)))

        # Value labels from first variable (they all share the same)
        vlabels = get_value_labels(var_names[0], meta)

        for val in all_vals:
            label = label_for_value(val, vlabels)
            # Count respondents who have this value in ANY of the variables
            mask = pd.Series(False, index=df.index)
            for vname in var_names:
                mask = mask | (df[vname] == val)
            if use_weight:
                n = float(df.loc[mask, weight_col].sum())
            else:
                n = int(mask.sum())
            pct = round(float(n) / float(total_cases) * 100, 1) if total_cases > 0 else 0.0
            rows.append((str(label), round(float(n), 1), pct))
    else:
        # MDGROUP ('d'): rows = variables, labels = variable labels
        for vname in var_names:
            label = get_var_label(vname, meta)
            mask = df[vname].notna() & (df[vname] != 0)
            if use_weight:
                n = float(df.loc[mask, weight_col].sum())
            else:
                n = int(mask.sum())
            pct = round(float(n) / float(total_cases) * 100, 1) if total_cases > 0 else 0.0
            rows.append((str(label), round(float(n), 1), pct))

    rows.append(('Total*', round(float(total_cases), 1), ''))

    return {
        'header': ['', 'n', '%'],
        'rows': rows,
        'caption': '*Multioption',
    }


def make_numeric_table(df, var_string, meta, col_map, full_stats=True, weight_col=None):
    """
    Tip 'n': Descriptive statistics (Mean, StdDev, Median, Min, Max, N)
    Tip 'm': Samo Mean i N

    Replicira: TABLES /OBSERVATION var1 var2 ... /TABLES (var1+var2+...)
               BY (STATISTICS) /STATISTICS mean stddev median minimum maximum validn
    """
    var_names_raw = parse_numeric_vars(var_string)
    var_names = [resolve_col(v, df, col_map) for v in var_names_raw]
    use_weight = weight_col and weight_col in df.columns

    if full_stats:
        header = ['', 'Mean', 'Std.Dev.', 'Median', 'Min', 'Max', 'N']
    else:
        header = ['', 'Mean', 'N']

    rows = []
    for vname in var_names:
        label = get_var_label(vname, meta)
        vals = pd.to_numeric(df[vname], errors='coerce')
        valid = vals.notna()
        col = vals[valid]

        if len(col) == 0:
            if full_stats:
                rows.append((str(label), '.', '.', '.', '.', '.', 0))
            else:
                rows.append((str(label), '.', 0))
        else:
            if use_weight:
                w = df.loc[valid, weight_col].values
                v = col.values
                w_sum = w.sum()
                w_mean = float(np.average(v, weights=w))
                w_n = round(float(w_sum), 1)

                if full_stats:
                    w_var = float(np.average((v - w_mean) ** 2, weights=w))
                    # Bessel correction za weighted std
                    w_std = float(np.sqrt(w_var * w_sum / (w_sum - 1))) if w_sum > 1 else 0.0
                    # Weighted median
                    sorted_idx = np.argsort(v)
                    sv, sw = v[sorted_idx], w[sorted_idx]
                    cum_w = np.cumsum(sw)
                    half = w_sum / 2.0
                    med_idx = np.searchsorted(cum_w, half)
                    w_med = float(sv[min(med_idx, len(sv) - 1)])

                    rows.append((str(label), round(w_mean, 2), round(w_std, 2),
                                 round(w_med, 2), round(float(v.min()), 2),
                                 round(float(v.max()), 2), w_n))
                else:
                    rows.append((str(label), round(w_mean, 2), w_n))
            else:
                mean_val = round(float(col.mean()), 2)
                n = int(len(col))

                if full_stats:
                    std_val = round(float(col.std(ddof=1)), 2)
                    med_val = round(float(col.median()), 2)
                    min_val = round(float(col.min()), 2)
                    max_val = round(float(col.max()), 2)
                    rows.append((str(label), mean_val, std_val, med_val, min_val, max_val, n))
                else:
                    rows.append((str(label), mean_val, n))

    return {
        'header': header,
        'rows': rows,
        'caption': '',
    }


def make_freq_table(df, var_name, meta, col_map, weight_col=None):
    """
    Tip 'f': FREQUENCIES procedura, sortirano po frekvenciji silazno.
    Replicira: FREQUENCIES VARIABLES=var /FORMAT=DFREQ /ORDER ANALYSIS.
    """
    actual_col = resolve_col(var_name.strip(), df, col_map)
    subset = df[df[actual_col].notna()].copy()
    val_labels = get_value_labels(actual_col, meta)
    use_weight = weight_col and weight_col in subset.columns

    if use_weight:
        counts = subset.groupby(actual_col)[weight_col].sum().sort_values(ascending=False)
    else:
        counts = subset[actual_col].value_counts()  # vec sortirano descending

    total_n = float(counts.sum())

    rows = []
    cum_pct = 0.0
    for val, n in counts.items():
        lbl = label_for_value(val, val_labels) if val_labels else label_for_value(val, {})
        n_val = float(n)
        pct = round(n_val / total_n * 100, 1) if total_n > 0 else 0.0
        cum_pct = round(cum_pct + pct, 1)
        rows.append((str(lbl), round(n_val, 1), pct, cum_pct))

    rows.append(('Total', round(total_n, 1), 100.0, 100.0))

    return {
        'header': ['', 'Frequency', '%', 'Cum. %'],
        'rows': rows,
        'caption': '',
    }


# ═══════════════════════════════════════════════════════════════════
#  EXCEL PISANJE
# ═══════════════════════════════════════════════════════════════════

def write_tables_to_excel(tables, output_path):
    """Pise sve tablice u Excel fajl s formatiranjem."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Total Tables"

    # Stilovi
    title_font = Font(name='Arial', size=11, bold=True)
    header_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    data_font = Font(name='Arial', size=10)
    total_font = Font(name='Arial', size=10, bold=True)
    caption_font = Font(name='Arial', size=9, italic=True, color='666666')
    pct_font = Font(name='Arial', size=10, color='333333')

    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    total_fill = PatternFill(start_color='D6DCE4', end_color='D6DCE4', fill_type='solid')
    even_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

    thin_side = Side(style='thin', color='B4B4B4')
    thin_border = Border(
        left=thin_side, right=thin_side,
        top=thin_side, bottom=thin_side,
    )

    # Track max widths per column
    col_widths = {}

    row_num = 1

    for table in tables:
        title_str = table['title']
        header = table['header']
        data_rows = table['rows']
        caption = table.get('caption', '')
        num_cols = len(header)

        # ── Naslov tablice ──
        cell = ws.cell(row=row_num, column=1, value=title_str)
        cell.font = title_font
        ws.merge_cells(
            start_row=row_num, start_column=1,
            end_row=row_num, end_column=num_cols,
        )
        row_num += 1

        # ── Header red ──
        for col_idx, h in enumerate(header, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(
                horizontal='center' if col_idx > 1 else 'left',
                vertical='center',
            )
            _track_width(col_widths, col_idx, h)
        row_num += 1

        # ── Data redovi ──
        for i, data_row in enumerate(data_rows):
            is_total = (i == len(data_rows) - 1)
            is_even = (i % 2 == 1) and not is_total

            for col_idx, val in enumerate(data_row, 1):
                cell = ws.cell(row=row_num, column=col_idx)

                if val is None or val == '':
                    cell.value = ''
                elif isinstance(val, str) and val == '.':
                    cell.value = '.'
                elif isinstance(val, float):
                    cell.value = val
                    cell.number_format = '0.0'
                elif isinstance(val, (int, np.integer)):
                    cell.value = int(val)
                    cell.number_format = '#,##0'
                else:
                    cell.value = str(val)

                cell.font = total_font if is_total else data_font
                cell.border = thin_border

                if is_total:
                    cell.fill = total_fill
                elif is_even:
                    cell.fill = even_fill

                if col_idx > 1:
                    cell.alignment = Alignment(horizontal='right')

                _track_width(col_widths, col_idx, val)

            row_num += 1

        # ── Caption ──
        if caption:
            cell = ws.cell(row=row_num, column=1, value=caption)
            cell.font = caption_font
            row_num += 1

        # Prazan red izmedju tablica
        row_num += 1

    # ── Postavi sirine kolona ──
    for col_idx, width in col_widths.items():
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = min(width + 3, 65)

    # Zamrzni header
    ws.freeze_panes = 'A2'

    wb.save(output_path)
    return len(tables)


def _track_width(col_widths, col_idx, value):
    """Prati maksimalnu sirinu kolone."""
    if value is None or value == '':
        length = 2
    else:
        length = len(str(value))
    current = col_widths.get(col_idx, 8)
    col_widths[col_idx] = max(current, length)


# ═══════════════════════════════════════════════════════════════════
#  KRIZANJA (CROSS-TABULATIONS) S TESTOM ZNACAJNOSTI
# ═══════════════════════════════════════════════════════════════════

SIG_LETTERS = list('ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz')
SIG_LEVEL = 1.96  # 95% confidence


def _col_pct_sig(p1, p2, n1, n2):
    """
    z-test za razliku dviju proporcija (column %).
    Isti izracun kao u 5_0_SEplusSig.SBS.
    p1, p2 su postoci (0-100), n1, n2 su baze.
    Vraca z-vrijednost ili 0 ako racunanje nije moguce.
    """
    if n1 < 30 or n2 < 30:
        return 0.0
    pooled = (n1 * p1 + n2 * p2) / (n1 + n2)
    denom_inner = (n1 * p1 + n2 * p2) / (n1 * n2) * (100 - pooled)
    if denom_inner <= 0:
        return 0.0
    return abs(p1 - p2) / np.sqrt(denom_inner)


def _mean_sig(m1, m2, sd1, sd2, n1, n2):
    """
    z-test za razliku dviju sredina (independent samples).
    Isti izracun kao u 5_0_SEplusSig.SBS za numericke tablice.
    """
    if n1 < 30 or n2 < 30:
        return 0.0
    v1 = sd1 ** 2
    v2 = sd2 ** 2
    denom = v1 / n1 + v2 / n2
    if denom <= 0:
        return 0.0
    return abs(m1 - m2) / np.sqrt(denom)


def make_crosstab_simple(df, var_name, break_var, meta, col_map, weight_col=None):
    """
    Krizanje za tip 's': pitanje BY break_var.
    Vraca dict s:
      col_labels: [labele kategorija break varijable + Total]
      col_letters: [A, B, C, ...]
      col_ns: [N za svaki stupac]
      row_labels: [labele kategorija pitanja]
      pct_matrix: 2D lista [row][col] postotaka
      sig_matrix: 2D lista [row][col] stringova slova znacajnosti
    """
    actual_col = resolve_col(var_name, df, col_map)
    actual_break = resolve_col(break_var, df, col_map)
    val_labels = get_value_labels(actual_col, meta)
    break_labels = get_value_labels(actual_break, meta)

    # Filtriraj valide na pitanju
    subset = df[df[actual_col].notna()].copy()
    use_weight = weight_col and weight_col in subset.columns

    # Kategorije break varijable (sortirane)
    break_vals = sorted(subset[actual_break].dropna().unique(),
                        key=lambda x: (0, float(x)) if isinstance(x, (int, float)) else (1, str(x)))

    # Kategorije pitanja (sortirane)
    all_values = sorted(subset[actual_col].dropna().unique(),
                        key=lambda x: (0, float(x)) if isinstance(x, (int, float)) else (1, str(x)))

    # Kolone: svaka kategorija break varijable + Total
    col_labels = [label_for_value(bv, break_labels) for bv in break_vals] + ['Total']
    col_letters = SIG_LETTERS[:len(break_vals)]  # bez Total-a
    n_cols = len(break_vals) + 1  # +1 za Total

    # N za svaki stupac
    col_ns = []
    for bv in break_vals:
        mask = subset[actual_break] == bv
        if use_weight:
            col_ns.append(float(subset.loc[mask, weight_col].sum()))
        else:
            col_ns.append(int(mask.sum()))
    # Total N
    if use_weight:
        col_ns.append(float(subset[weight_col].sum()))
    else:
        col_ns.append(len(subset))

    # Racunaj postotke
    row_labels = [label_for_value(v, val_labels) for v in all_values]
    pct_matrix = []
    for val in all_values:
        row_pcts = []
        for ci, bv in enumerate(break_vals):
            mask = (subset[actual_col] == val) & (subset[actual_break] == bv)
            if use_weight:
                n = float(subset.loc[mask, weight_col].sum())
            else:
                n = int(mask.sum())
            pct = round(n / col_ns[ci] * 100, 1) if col_ns[ci] > 0 else 0.0
            row_pcts.append(pct)
        # Total %
        mask_total = subset[actual_col] == val
        if use_weight:
            n_total = float(subset.loc[mask_total, weight_col].sum())
        else:
            n_total = int(mask_total.sum())
        pct_total = round(n_total / col_ns[-1] * 100, 1) if col_ns[-1] > 0 else 0.0
        row_pcts.append(pct_total)
        pct_matrix.append(row_pcts)

    # Test znacajnosti (parovi stupaca unutar break varijable, bez Total-a)
    sig_matrix = _compute_sig_pct(pct_matrix, col_ns, col_letters, len(break_vals))

    return {
        'type': 'simple',
        'col_labels': col_labels,
        'col_letters': col_letters,
        'col_ns': col_ns,
        'row_labels': row_labels,
        'pct_matrix': pct_matrix,
        'sig_matrix': sig_matrix,
    }


def make_crosstab_mr(df, var_string, break_var, meta, col_map, mr_type='k', weight_col=None):
    """Krizanje za tip 'k'/'d': MR pitanje BY break_var."""
    var_names_raw = parse_mr_vars(var_string)
    var_names = [resolve_col(v, df, col_map) for v in var_names_raw]
    actual_break = resolve_col(break_var, df, col_map)
    break_labels = get_value_labels(actual_break, meta)

    mr_df = df[var_names]
    valid_mask = mr_df.notna().any(axis=1)
    subset = df[valid_mask].copy()
    use_weight = weight_col and weight_col in subset.columns

    break_vals = sorted(subset[actual_break].dropna().unique(),
                        key=lambda x: (0, float(x)) if isinstance(x, (int, float)) else (1, str(x)))

    col_labels = [label_for_value(bv, break_labels) for bv in break_vals] + ['Total']
    col_letters = SIG_LETTERS[:len(break_vals)]

    # N za svaki stupac (cases, ne responses)
    col_ns = []
    for bv in break_vals:
        mask = subset[actual_break] == bv
        if use_weight:
            col_ns.append(float(subset.loc[mask, weight_col].sum()))
        else:
            col_ns.append(int(mask.sum()))
    if use_weight:
        col_ns.append(float(subset[weight_col].sum()))
    else:
        col_ns.append(len(subset))

    if mr_type == 'k':
        # MRGROUP: rows = unique values across all variables
        all_vals = set()
        for vname in var_names:
            all_vals.update(subset[vname].dropna().unique())
        all_vals.discard(0)
        all_vals = sorted(all_vals,
                          key=lambda x: (0, float(x)) if isinstance(x, (int, float)) else (1, str(x)))
        vlabels = get_value_labels(var_names[0], meta)
        row_labels = [label_for_value(val, vlabels) for val in all_vals]

        pct_matrix = []
        for val in all_vals:
            row_pcts = []
            # Respondent has this value in ANY of the variables
            item_mask = pd.Series(False, index=subset.index)
            for vname in var_names:
                item_mask = item_mask | (subset[vname] == val)
            for ci, bv in enumerate(break_vals):
                combined = item_mask & (subset[actual_break] == bv)
                if use_weight:
                    n = float(subset.loc[combined, weight_col].sum())
                else:
                    n = int(combined.sum())
                pct = round(n / col_ns[ci] * 100, 1) if col_ns[ci] > 0 else 0.0
                row_pcts.append(pct)
            # Total
            if use_weight:
                n_total = float(subset.loc[item_mask, weight_col].sum())
            else:
                n_total = int(item_mask.sum())
            pct_total = round(n_total / col_ns[-1] * 100, 1) if col_ns[-1] > 0 else 0.0
            row_pcts.append(pct_total)
            pct_matrix.append(row_pcts)
    else:
        # MDGROUP ('d'): rows = variables, labels = variable labels
        row_labels = [get_var_label(v, meta) for v in var_names]
        pct_matrix = []
        for vname in var_names:
            row_pcts = []
            item_mask = subset[vname].notna() & (subset[vname] != 0)
            for ci, bv in enumerate(break_vals):
                combined = item_mask & (subset[actual_break] == bv)
                if use_weight:
                    n = float(subset.loc[combined, weight_col].sum())
                else:
                    n = int(combined.sum())
                pct = round(n / col_ns[ci] * 100, 1) if col_ns[ci] > 0 else 0.0
                row_pcts.append(pct)
            # Total
            if use_weight:
                n_total = float(subset.loc[item_mask, weight_col].sum())
            else:
                n_total = int(item_mask.sum())
            pct_total = round(n_total / col_ns[-1] * 100, 1) if col_ns[-1] > 0 else 0.0
            row_pcts.append(pct_total)
            pct_matrix.append(row_pcts)

    sig_matrix = _compute_sig_pct(pct_matrix, col_ns, col_letters, len(break_vals))

    return {
        'type': 'mr',
        'col_labels': col_labels,
        'col_letters': col_letters,
        'col_ns': col_ns,
        'row_labels': row_labels,
        'pct_matrix': pct_matrix,
        'sig_matrix': sig_matrix,
        'caption': '*Multioption',
    }


def make_crosstab_numeric(df, var_string, break_var, meta, col_map,
                          full_stats=True, weight_col=None):
    """Krizanje za tip 'n'/'m': numericka pitanja BY break_var."""
    var_names_raw = parse_numeric_vars(var_string)
    var_names = [resolve_col(v, df, col_map) for v in var_names_raw]
    actual_break = resolve_col(break_var, df, col_map)
    break_labels = get_value_labels(actual_break, meta)

    use_weight = weight_col and weight_col in df.columns

    break_vals = sorted(df[actual_break].dropna().unique(),
                        key=lambda x: (0, float(x)) if isinstance(x, (int, float)) else (1, str(x)))

    col_labels = [label_for_value(bv, break_labels) for bv in break_vals] + ['Total']
    col_letters = SIG_LETTERS[:len(break_vals)]

    # Za svaku varijablu: mean po svakom stupcu
    row_labels = [get_var_label(v, meta) for v in var_names]
    mean_matrix = []  # [row][col]
    sd_matrix = []    # za sig test
    n_matrix = []     # N po celiji

    for vname in var_names:
        vals = pd.to_numeric(df[vname], errors='coerce')

        row_means = []
        row_sds = []
        row_ns = []

        for bv in break_vals:
            mask = vals.notna() & (df[actual_break] == bv)
            col_vals = vals[mask]
            if use_weight:
                w = df.loc[mask, weight_col].values
                if len(col_vals) > 0 and w.sum() > 0:
                    m = float(np.average(col_vals.values, weights=w))
                    sd = float(np.sqrt(np.average((col_vals.values - m) ** 2, weights=w)))
                    n = float(w.sum())
                else:
                    m, sd, n = 0, 0, 0
            else:
                if len(col_vals) > 0:
                    m = float(col_vals.mean())
                    sd = float(col_vals.std(ddof=1)) if len(col_vals) > 1 else 0.0
                    n = len(col_vals)
                else:
                    m, sd, n = 0, 0, 0
            row_means.append(round(m, 2))
            row_sds.append(round(sd, 2))
            row_ns.append(round(n, 1) if use_weight else n)

        # Total
        mask_total = vals.notna()
        col_vals_total = vals[mask_total]
        if use_weight:
            w = df.loc[mask_total, weight_col].values
            if len(col_vals_total) > 0 and w.sum() > 0:
                m = float(np.average(col_vals_total.values, weights=w))
                sd = float(np.sqrt(np.average((col_vals_total.values - m) ** 2, weights=w)))
                n = float(w.sum())
            else:
                m, sd, n = 0, 0, 0
        else:
            if len(col_vals_total) > 0:
                m = float(col_vals_total.mean())
                sd = float(col_vals_total.std(ddof=1)) if len(col_vals_total) > 1 else 0.0
                n = len(col_vals_total)
            else:
                m, sd, n = 0, 0, 0
        row_means.append(round(m, 2))
        row_sds.append(round(sd, 2))
        row_ns.append(round(n, 1) if use_weight else n)

        mean_matrix.append(row_means)
        sd_matrix.append(row_sds)
        n_matrix.append(row_ns)

    # Sig test za meanove
    sig_matrix = []
    num_break = len(break_vals)
    for ri in range(len(var_names)):
        sig_row = []
        for ci in range(num_break + 1):
            if ci >= num_break:  # Total — nema sig testiranja
                sig_row.append('')
                continue
            letters = ''
            for oi in range(num_break):
                if oi == ci:
                    continue
                z = _mean_sig(
                    mean_matrix[ri][ci], mean_matrix[ri][oi],
                    sd_matrix[ri][ci], sd_matrix[ri][oi],
                    n_matrix[ri][ci], n_matrix[ri][oi],
                )
                if z >= SIG_LEVEL and mean_matrix[ri][ci] > mean_matrix[ri][oi]:
                    letters += col_letters[oi]
            sig_row.append(letters)
        sig_matrix.append(sig_row)

    # col_ns = N za prvi red (ili agregirano)
    col_ns = n_matrix[0] if n_matrix else [0] * (num_break + 1)

    return {
        'type': 'numeric',
        'col_labels': col_labels,
        'col_letters': col_letters,
        'col_ns': col_ns,
        'row_labels': row_labels,
        'mean_matrix': mean_matrix,
        'sd_matrix': sd_matrix,
        'n_matrix': n_matrix,
        'sig_matrix': sig_matrix,
    }


def _compute_sig_pct(pct_matrix, col_ns, col_letters, num_break):
    """Izracunaj znacajnost za matricu postotaka. Vraca sig_matrix."""
    sig_matrix = []
    for row_pcts in pct_matrix:
        sig_row = []
        for ci in range(num_break + 1):
            if ci >= num_break:  # Total — nema sig testiranja
                sig_row.append('')
                continue
            letters = ''
            for oi in range(num_break):
                if oi == ci:
                    continue
                z = _col_pct_sig(row_pcts[ci], row_pcts[oi], col_ns[ci], col_ns[oi])
                if z >= SIG_LEVEL and row_pcts[ci] > row_pcts[oi]:
                    letters += col_letters[oi]
            sig_row.append(letters)
        sig_matrix.append(sig_row)
    return sig_matrix


# ═══════════════════════════════════════════════════════════════════
#  BANNER: Merge multiple crosstabs into one wide table
# ═══════════════════════════════════════════════════════════════════

# Alternate header fills for visual group separation
_BANNER_FILLS = [
    PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid'),  # blue
    PatternFill(start_color='548235', end_color='548235', fill_type='solid'),  # green
    PatternFill(start_color='BF8F00', end_color='BF8F00', fill_type='solid'),  # gold
    PatternFill(start_color='C55A11', end_color='C55A11', fill_type='solid'),  # orange
    PatternFill(start_color='7030A0', end_color='7030A0', fill_type='solid'),  # purple
    PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid'),  # light blue
]


def merge_crosstabs_banner(crosstabs):
    """
    Merge multiple single-break-var crosstab results into one banner structure.
    Each crosstab dict has col_labels (incl Total), col_letters, col_ns, row_labels,
    pct_matrix/mean_matrix, sig_matrix.
    Returns a 'banner' dict with 'groups' list for the writer.
    """
    if not crosstabs:
        return None

    first = crosstabs[0]
    is_numeric = first['type'] == 'numeric'
    row_labels = first['row_labels']

    groups = []
    for xt in crosstabs:
        # Strip Total from col_labels / col_ns (last element)
        n_break = len(xt['col_letters'])  # number of break categories (excludes Total)
        grp = {
            'col_labels': xt['col_labels'][:n_break],
            'col_letters': xt['col_letters'],
            'col_ns': xt['col_ns'][:n_break],
        }
        if is_numeric:
            # Strip Total column from each row
            grp['mean_matrix'] = [row[:n_break] for row in xt['mean_matrix']]
            grp['sd_matrix'] = [row[:n_break] for row in xt.get('sd_matrix', [])]
            grp['n_matrix'] = [row[:n_break] for row in xt.get('n_matrix', [])]
        else:
            grp['pct_matrix'] = [row[:n_break] for row in xt['pct_matrix']]
        grp['sig_matrix'] = [row[:n_break] for row in xt['sig_matrix']]
        groups.append(grp)

    # Total column from first crosstab (all should have the same Total)
    first_n_break = len(first['col_letters'])
    total_n = first['col_ns'][first_n_break]  # last element
    if is_numeric:
        total_means = [row[first_n_break] for row in first['mean_matrix']]
        total_sds = [row[first_n_break] for row in first.get('sd_matrix', [])]
    else:
        total_pcts = [row[first_n_break] for row in first['pct_matrix']]

    banner = {
        'type': 'numeric' if is_numeric else 'pct',
        'row_labels': row_labels,
        'groups': groups,
        'total_n': total_n,
        'caption': first.get('caption', ''),
    }
    if is_numeric:
        banner['total_means'] = total_means
        banner['total_sds'] = total_sds
    else:
        banner['total_pcts'] = total_pcts

    return banner


def _make_global_letters(groups):
    """Create continuous letter sequence across all banner groups for sig_total.
    Returns: list of letters (one per column across all groups),
             and list of (letter, label) tuples for legend.
    """
    letters = []
    legend = []
    idx = 0
    for grp in groups:
        for ci, lbl in enumerate(grp['col_labels']):
            letter = SIG_LETTERS[idx] if idx < len(SIG_LETTERS) else f'A{idx}'
            letters.append(letter)
            legend.append((letter, lbl))
            idx += 1
    return letters, legend


def compute_sig_total_banner(banner):
    """
    Compute significance of Total column vs each banner category column.
    Same z-test as 5_1_Sig_total.SBS.
    Uses continuous letters across ALL groups so Total column letters are unambiguous.
    Returns dict with 'letters' (list of strings per row),
    'directions' (list of dicts: global_letter -> '↑'/'↓'),
    'global_letters'/'legend' for header/legend rendering.
    """
    groups = banner['groups']
    is_numeric = banner['type'] == 'numeric'
    total_n = banner['total_n']
    n_rows = len(banner['row_labels'])

    global_letters, legend = _make_global_letters(groups)

    sig_total = []
    directions = []  # per row: {global_letter: '↑' or '↓'}
    flat_idx = 0  # running index across all groups/columns
    for ri in range(n_rows):
        letters = ''
        row_dirs = {}
        flat_idx = 0
        if is_numeric:
            total_val = banner['total_means'][ri]
            total_sd = banner.get('total_sds', [0] * n_rows)[ri]
        else:
            total_val = banner['total_pcts'][ri]

        for grp in groups:
            n_cols = len(grp['col_labels'])
            for ci in range(n_cols):
                col_n = grp['col_ns'][ci]
                gl = global_letters[flat_idx]

                if is_numeric:
                    col_val = grp['mean_matrix'][ri][ci]
                    col_sd = grp.get('sd_matrix', [[0]*n_cols]*n_rows)[ri][ci]
                    col_n_i = grp.get('n_matrix', [[0]*n_cols]*n_rows)[ri][ci]
                    z = _mean_sig(total_val, col_val, total_sd, col_sd,
                                  total_n, col_n_i if col_n_i else col_n)
                else:
                    col_val = grp['pct_matrix'][ri][ci]
                    z = _col_pct_sig(total_val, col_val, total_n, col_n)

                if z >= SIG_LEVEL:
                    letters += gl
                    row_dirs[gl] = '↑' if col_val > total_val else '↓'
                flat_idx += 1

        sig_total.append(letters)
        directions.append(row_dirs)

    return {'letters': sig_total, 'directions': directions,
            'global_letters': global_letters, 'legend': legend}


def write_banner_to_sheet(ws, banner, title_str, start_row=1, show_sig=True,
                          show_sig_total=False):
    """
    Write a merged banner crosstab to an Excel sheet.
    All banner groups side by side, with visual separation, one Total column.
    show_sig_total: if True, Total column gets sig letters vs each category.
    Returns the next free row number.
    """
    title_font = Font(name='Arial', size=11, bold=True)
    header_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    data_font = Font(name='Arial', size=10)
    base_font = Font(name='Arial', size=9, italic=True, color='666666')
    letter_font = Font(name='Arial', size=8, bold=True, color='4472C4')

    total_fill = PatternFill(start_color='D6DCE4', end_color='D6DCE4', fill_type='solid')
    even_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    sig_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')

    thick_side = Side(style='medium', color='333333')
    thin_side = Side(style='thin', color='B4B4B4')
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    groups = banner['groups']
    row_labels = banner['row_labels']
    is_numeric = banner['type'] == 'numeric'

    # Calculate total number of data columns = sum of group cols + 1 (Total)
    total_data_cols = sum(len(g['col_labels']) for g in groups) + 1
    n_total_cols = 1 + total_data_cols  # +1 for label column

    row_num = start_row

    # ── Title row ──
    cell = ws.cell(row=row_num, column=1, value=title_str)
    cell.font = title_font
    ws.merge_cells(start_row=row_num, start_column=1,
                   end_row=row_num, end_column=n_total_cols)
    row_num += 1

    # ── Header row: category labels ──
    ws.cell(row=row_num, column=1, value='').border = thin_border
    col_offset = 2  # first data column
    for gi, grp in enumerate(groups):
        fill = _BANNER_FILLS[gi % len(_BANNER_FILLS)]
        for ci, lbl in enumerate(grp['col_labels']):
            cell = ws.cell(row=row_num, column=col_offset + ci, value=lbl)
            cell.font = header_font
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
            # Thick left border on first column of each group (visual separator)
            if ci == 0 and gi > 0:
                cell.border = Border(left=thick_side, right=thin_side,
                                     top=thin_side, bottom=thin_side)
        col_offset += len(grp['col_labels'])
    # Total column
    cell = ws.cell(row=row_num, column=col_offset, value='Total')
    cell.font = Font(name='Arial', size=10, bold=True)
    cell.fill = total_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center')
    row_num += 1

    # ── Pre-compute sig_total if needed ──
    sig_total_result = None
    sig_total_letters = None
    if show_sig_total:
        sig_total_result = compute_sig_total_banner(banner)
        sig_total_letters = sig_total_result['letters']

    # ── Letter row (A, B, C... restart per group) — only if show_sig ──
    # For sig_total sheet, use continuous global letters instead of per-group
    if show_sig:
        # Determine which letters to display per column
        if show_sig_total and sig_total_result:
            _gl = sig_total_result['global_letters']
        else:
            _gl = None

        ws.cell(row=row_num, column=1, value='').border = thin_border
        col_offset = 2
        flat_idx = 0
        for gi, grp in enumerate(groups):
            for ci, _per_grp_letter in enumerate(grp['col_letters']):
                n = grp['col_ns'][ci]
                base_mark = ''
                if n < 30:
                    base_mark = '**'
                elif n < 50:
                    base_mark = '*'
                display_letter = _gl[flat_idx] if _gl else _per_grp_letter
                cell = ws.cell(row=row_num, column=col_offset + ci,
                               value=f"{display_letter}{base_mark}")
                cell.font = letter_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                if ci == 0 and gi > 0:
                    cell.border = Border(left=thick_side, right=thin_side,
                                         top=thin_side, bottom=thin_side)
                flat_idx += 1
            col_offset += len(grp['col_labels'])
        # Total — empty
        ws.cell(row=row_num, column=col_offset, value='').border = thin_border
        row_num += 1

    # ── Data rows ──
    for ri, label in enumerate(row_labels):
        is_even = (ri % 2 == 1)
        cell = ws.cell(row=row_num, column=1, value=label)
        cell.font = data_font
        cell.border = thin_border
        if is_even:
            cell.fill = even_fill

        col_offset = 2
        flat_idx = 0
        for gi, grp in enumerate(groups):
            n_gcols = len(grp['col_labels'])
            for ci in range(n_gcols):
                cell = ws.cell(row=row_num, column=col_offset + ci)
                brd = thin_border
                if ci == 0 and gi > 0:
                    brd = Border(left=thick_side, right=thin_side,
                                 top=thin_side, bottom=thin_side)
                cell.border = brd

                # Check if this cell is sig different from Total
                _st_hit = False
                _st_arrow = ''
                if show_sig_total and sig_total_result:
                    gl = sig_total_result['global_letters'][flat_idx]
                    if gl in (sig_total_letters[ri] if sig_total_letters else ''):
                        _st_hit = True
                        _st_arrow = sig_total_result['directions'][ri].get(gl, '')

                if is_numeric:
                    val = grp['mean_matrix'][ri][ci]
                    sig = grp['sig_matrix'][ri][ci] if (show_sig and not show_sig_total and ri < len(grp['sig_matrix'])) else ''
                    if val == 0 and grp.get('n_matrix', [[]])[ri][ci] == 0:
                        cell.value = '.'
                    elif sig:
                        cell.value = f"{val:.2f}\n{sig}"
                        cell.fill = sig_fill
                        cell.alignment = Alignment(horizontal='right', wrap_text=True)
                    elif _st_hit:
                        cell.value = f"{val:.2f}\n{_st_arrow}"
                        cell.fill = sig_fill
                        cell.alignment = Alignment(horizontal='right', wrap_text=True)
                    else:
                        cell.value = val
                        cell.number_format = '0.00'
                        cell.alignment = Alignment(horizontal='right')
                else:
                    pct = grp['pct_matrix'][ri][ci]
                    sig = grp['sig_matrix'][ri][ci] if (show_sig and not show_sig_total and ri < len(grp['sig_matrix'])) else ''
                    if pct == 0 and not sig:
                        cell.value = 0.0
                        cell.number_format = '0.0'
                        cell.alignment = Alignment(horizontal='right')
                        if _st_hit:
                            cell.value = f"0.0\n{_st_arrow}"
                            cell.fill = sig_fill
                            cell.alignment = Alignment(horizontal='right', wrap_text=True)
                    elif sig:
                        cell.value = f"{pct:.1f}\n{sig}"
                        cell.fill = sig_fill
                        cell.alignment = Alignment(horizontal='right', wrap_text=True)
                    elif _st_hit:
                        cell.value = f"{pct:.1f}\n{_st_arrow}"
                        cell.fill = sig_fill
                        cell.alignment = Alignment(horizontal='right', wrap_text=True)
                    else:
                        cell.value = pct
                        cell.number_format = '0.0'
                        cell.alignment = Alignment(horizontal='right')

                cell.font = data_font
                flat_idx += 1
                if is_even and not _st_hit and not (show_sig and
                        (grp['sig_matrix'][ri][ci] if ri < len(grp['sig_matrix']) else '')):
                    cell.fill = even_fill

            col_offset += n_gcols

        # Total column
        cell = ws.cell(row=row_num, column=col_offset)
        cell.border = thin_border
        cell.font = data_font
        st_letters = sig_total_letters[ri] if sig_total_letters else ''
        if is_numeric:
            val_t = banner['total_means'][ri]
            if st_letters:
                cell.value = f"{val_t:.2f}\n{st_letters}"
                cell.fill = sig_fill
                cell.alignment = Alignment(horizontal='right', wrap_text=True)
            else:
                cell.value = val_t
                cell.number_format = '0.00'
                cell.alignment = Alignment(horizontal='right')
        else:
            pct_t = banner['total_pcts'][ri]
            if pct_t == 0 and not st_letters:
                cell.value = 0.0
                cell.number_format = '0.0'
                cell.alignment = Alignment(horizontal='right')
            elif st_letters:
                cell.value = f"{pct_t:.1f}\n{st_letters}"
                cell.fill = sig_fill
                cell.alignment = Alignment(horizontal='right', wrap_text=True)
            else:
                cell.value = pct_t
                cell.number_format = '0.0'
                cell.alignment = Alignment(horizontal='right')
        if is_even and not st_letters:
            cell.fill = even_fill

        row_num += 1

    # ── N row ──
    cell = ws.cell(row=row_num, column=1, value='N')
    cell.font = base_font
    cell.border = thin_border
    cell.fill = total_fill
    col_offset = 2
    for gi, grp in enumerate(groups):
        for ci, n in enumerate(grp['col_ns']):
            cell = ws.cell(row=row_num, column=col_offset + ci, value=round(n))
            cell.font = base_font
            cell.border = thin_border
            cell.fill = total_fill
            cell.alignment = Alignment(horizontal='right')
            if ci == 0 and gi > 0:
                cell.border = Border(left=thick_side, right=thin_side,
                                     top=thin_side, bottom=thin_side)
        col_offset += len(grp['col_labels'])
    # Total N
    cell = ws.cell(row=row_num, column=col_offset, value=round(banner['total_n']))
    cell.font = base_font
    cell.border = thin_border
    cell.fill = total_fill
    cell.alignment = Alignment(horizontal='right')
    row_num += 1

    # ── Caption ──
    caption = banner.get('caption', '')
    if caption:
        cell = ws.cell(row=row_num, column=1, value=caption)
        cell.font = Font(name='Arial', size=9, italic=True, color='666666')
        row_num += 1

    # ── Legend (only if show_sig) ──
    if show_sig:
        row_num += 1
        ws.cell(row=row_num, column=1, value='* small base (30≤n<50)').font = \
            Font(name='Arial', size=8, italic=True, color='999999')
        row_num += 1
        ws.cell(row=row_num, column=1, value='** very small base (n<30) — ineligible for sig testing').font = \
            Font(name='Arial', size=8, italic=True, color='999999')
        row_num += 1
        ws.cell(row=row_num, column=1, value='Significance test: z-test, 95% confidence').font = \
            Font(name='Arial', size=8, italic=True, color='999999')

    # ── Column widths ──
    ws.column_dimensions['A'].width = 45
    for ci in range(total_data_cols):
        col_letter = get_column_letter(ci + 2)
        ws.column_dimensions[col_letter].width = 14

    if start_row == 1:
        ws.freeze_panes = 'B4' if show_sig else 'B3'

    return row_num


def main():
    parser = argparse.ArgumentParser(
        description='Generira Excel tablice iz SPSS .sav fajla koristeci input.txt definicije',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Primjeri:
  python spss_tables.py --sav data.sav --input input.txt
  python spss_tables.py --sav data.sav --input input.txt --output rezultati.xlsx --start 1

Potrebni paketi:
  pip install pyreadstat pandas openpyxl
        """,
    )
    parser.add_argument('--sav', required=True, help='Putanja do SPSS .sav fajla')
    parser.add_argument('--input', required=True, help='Putanja do input.txt definicijskog fajla')
    parser.add_argument('--output', default='output_tables.xlsx', help='Izlazni Excel fajl (default: output_tables.xlsx)')
    parser.add_argument('--start', type=int, default=1, help='Pocetni broj tablice / pomak (default: 1)')
    parser.add_argument('--weight', default=None, help='Ime varijable za ponderiranje (npr. ponder, wgt)')
    parser.add_argument('--filter', default=None, help='Ime filter varijable (zadrzava redove gdje var != 0 i nije missing)')

    args = parser.parse_args()

    # ── Provjeri fajlove ──
    if not os.path.exists(args.sav):
        print(f"GRESKA: SAV fajl ne postoji: {args.sav}")
        sys.exit(1)
    if not os.path.exists(args.input):
        print(f"GRESKA: Input fajl ne postoji: {args.input}")
        sys.exit(1)

    t_start = time.time()

    # ── Citaj .sav ──
    print(f"Citam podatke: {args.sav}")
    df, meta = pyreadstat.read_sav(args.sav, apply_value_formats=False)
    print(f"  -> {len(df)} redova, {len(df.columns)} varijabli")

    col_map = build_column_map(df)

    # ── Primijeni filter ──
    weight_col = None
    if args.filter:
        try:
            filt_col = resolve_col(args.filter, df, col_map)
            before = len(df)
            df = df[df[filt_col].notna() & (df[filt_col] != 0)].copy()
            col_map = build_column_map(df)
            print(f"  Filter '{filt_col}': {before} -> {len(df)} redova")
        except KeyError:
            print(f"  UPOZORENJE: Filter varijabla '{args.filter}' ne postoji, ignoriram")

    # ── Postavi weight ──
    if args.weight:
        try:
            weight_col = resolve_col(args.weight, df, col_map)
            w_sum = df[weight_col].sum()
            print(f"  Weight '{weight_col}': suma={w_sum:.1f}, N={len(df)}")
        except KeyError:
            print(f"  UPOZORENJE: Weight varijabla '{args.weight}' ne postoji, ignoriram")
            weight_col = None

    # ── Parsiraj input.txt ──
    print(f"Citam definicije: {args.input}")
    break_vars, titles, variables = parse_input_file(args.input)
    print(f"  -> {len(break_vars)} break varijabli, {len(titles)} naslova, {len(variables)} var definicija")

    count = min(len(titles), len(variables))
    if len(titles) != len(variables):
        print(f"  UPOZORENJE: Broj naslova ({len(titles)}) != broj varijabli ({len(variables)}), koristim {count}")

    pomak = args.start

    # ── Generiraj tablice ──
    tables = []
    errors = 0

    print(f"\nGeneriram tablice...")
    for i in range(count):
        title_line = titles[i]
        var_line = variables[i]

        table_type = get_table_type(title_line)
        table_title = get_table_title(title_line)
        table_num = i + pomak

        title_str = f"{table_title} (Table {table_num}.1)"

        try:
            if table_type == 's':
                result = make_simple_table(df, var_line.strip(), meta, col_map, weight_col)

            elif table_type in ('k', 'd'):
                result = make_mr_table(df, var_line, meta, col_map, table_type, weight_col)

            elif table_type == 'n':
                result = make_numeric_table(df, var_line, meta, col_map, full_stats=True, weight_col=weight_col)

            elif table_type == 'm':
                result = make_numeric_table(df, var_line, meta, col_map, full_stats=False, weight_col=weight_col)

            elif table_type == 'f':
                result = make_freq_table(df, var_line, meta, col_map, weight_col)

            else:
                print(f"  [{table_num}] NEPOZNAT tip '{table_type}': {table_title[:60]}")
                errors += 1
                continue

            result['title'] = title_str
            tables.append(result)
            print(f"  [{table_num:3d}] {table_type}: {table_title[:70]}")

        except KeyError as e:
            print(f"  [{table_num:3d}] GRESKA - varijabla {e} ne postoji u .sav fajlu")
            errors += 1
        except Exception as e:
            print(f"  [{table_num:3d}] GRESKA - {e}")
            errors += 1

    # ── Pisi u Excel ──
    print(f"\nPisem {len(tables)} tablica u: {args.output}")
    count_written = write_tables_to_excel(tables, args.output)

    elapsed = time.time() - t_start

    print(f"\n{'='*50}")
    print(f"Gotovo!")
    print(f"  Tablica generiranih: {count_written}")
    if errors > 0:
        print(f"  Gresaka:             {errors}")
    print(f"  Vrijeme:             {elapsed:.2f} sekundi")
    print(f"  Output:              {os.path.abspath(args.output)}")
    print(f"{'='*50}")


if __name__ == '__main__':
    main()
