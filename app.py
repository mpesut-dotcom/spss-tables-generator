#!/usr/bin/env python3
"""
SPSS Tables — Web aplikacija
=============================
Korisničko sučelje za generiranje Excel tablica iz SPSS .sav podataka.

Pokretanje:
    streamlit run app.py

Potrebni paketi:
    pip install streamlit pyreadstat pandas openpyxl
"""

import hashlib
import io
import json
import os
import re
import tempfile
import time

import numpy as np
import pandas as pd
import pyreadstat
import streamlit as st
from openpyxl import load_workbook


# Importaj engine funkcije iz spss_tables.py
from spo_parser import parse_spo, sav_names_match, parse_filter_expression, match_spo_to_input
from spss_tables import (
    build_column_map,
    compute_sig_total_banner,
    get_table_title,
    get_table_type,
    make_crosstab_mr,
    make_crosstab_numeric,
    make_crosstab_simple,
    make_freq_table,
    make_mr_table,
    make_numeric_table,
    make_simple_table,
    merge_crosstabs_banner,
    parse_numeric_vars,
    write_banner_to_sheet,
    write_tables_to_excel,
)


# ═══════════════════════════════════════════════════════════════════
#  PARSIRANJE INPUT-a IZ UPLOADANOG FAJLA (bytes, ne filepath)
# ═══════════════════════════════════════════════════════════════════

def parse_input_bytes(raw_bytes):
    """Parsira uploadani input.txt iz sirovih bajtova."""
    for enc in ('utf-8', 'cp1250', 'latin-1'):
        try:
            content = raw_bytes.decode(enc)
            break
        except (UnicodeDecodeError, AttributeError):
            continue
    else:
        content = raw_bytes.decode('utf-8', errors='replace')

    lines = content.replace('\r\n', '\n').split('\n')

    sections = []
    current = []
    for line in lines:
        if line.strip() == '':
            if current:
                sections.append(current)
                current = []
        else:
            current.append(line)
    if current:
        sections.append(current)

    if len(sections) < 3:
        raise ValueError(
            f"Input fajl mora imati 3 sekcije razdvojene blank linijama, "
            f"pronađeno: {len(sections)}"
        )

    return sections[0], sections[1], sections[2]


def validate_input(titles, variables, df_columns, break_vars, df=None, meta=None):
    """
    Validate input script against the data file.
    Returns list of warning dicts: {'level': 'error'|'warning'|'info', 'msg': str}
    """
    warnings_list = []
    col_set = {c.lower() for c in df_columns}

    # Pre-build data base→vars map for per-line completeness checks
    _q_var_pat = re.compile(r'^(q\d+|r\d+)_(\d+)$', re.I)
    _data_base_vars = {}
    for c in df_columns:
        if df is not None and c in df.columns and pd.api.types.is_object_dtype(df[c]):
            continue
        m = _q_var_pat.match(c)
        if m:
            _data_base_vars.setdefault(m.group(1).lower(), set()).add(c.lower())

    def _snippet(row_idx):
        """Build context snippet for a given row index."""
        parts = []
        if row_idx < len(titles):
            parts.append(f"Naslov:    {titles[row_idx].strip()}")
        if row_idx < len(variables):
            parts.append(f"Varijable: {variables[row_idx].strip()}")
        return '\n'.join(parts) if parts else None

    # 1. Check section lengths match
    if len(titles) != len(variables):
        diff = abs(len(titles) - len(variables))
        longer = "naslova" if len(titles) > len(variables) else "varijabli"
        shift_hint = ""
        shift_idx = None
        for idx in range(len(titles)):
            t = titles[idx].strip()
            if t.startswith('$') or ('+' in t and re.match(r'^\w+\s', t)):
                shift_hint = f" Red {idx+1} u sekciji naslova izgleda kao definicija varijabli."
                shift_idx = idx
                break
        if not shift_hint:
            for idx in range(len(variables)):
                v = variables[idx].strip()
                if re.match(r'^[skdnmf]\s', v):
                    shift_hint = f" Red {idx+1} u sekciji varijabli izgleda kao naslov tablice."
                    shift_idx = idx
                    break
        # Build a multi-line snippet showing the mismatch area
        shift_snippet = None
        if shift_idx is not None:
            lines = []
            for off in range(max(0, shift_idx - 1), min(max(len(titles), len(variables)), shift_idx + 3)):
                t_txt = titles[off].strip() if off < len(titles) else '(nema)'
                v_txt = variables[off].strip() if off < len(variables) else '(nema)'
                marker = '  ◄◄◄' if off == shift_idx else ''
                lines.append(f"Red {off+1}:{marker}\n  N: {t_txt}\n  V: {v_txt}")
            shift_snippet = '\n'.join(lines)
        warnings_list.append({
            'level': 'error',
            'msg': f"Broj naslova ({len(titles)}) ne odgovara broju varijabli ({len(variables)}) "
                   f"— {diff} {'više' if len(titles) > len(variables) else 'manje'} {longer}. "
                   f"Moguće pomaknut red čitanja u input skripti.{shift_hint}",
            'row': (shift_idx + 1) if shift_idx is not None else None,
            'snippet': shift_snippet,
        })
        # When rows are shifted, other checks are unreliable — return early
        return warnings_list

    count = min(len(titles), len(variables))

    # 2. Check break vars exist
    for bv in break_vars:
        if bv.strip().lower() not in col_set and bv.strip().lower() != 'id':
            warnings_list.append({
                'level': 'warning',
                'msg': f"Break varijabla '{bv.strip()}' iz input skripte ne postoji u datafileu.",
                'row': None, 'snippet': None,
            })

    # 3. Per-table checks
    missing_vars = []
    possible_cross_input = False

    for i in range(count):
        title_line = titles[i].strip()
        var_line = variables[i].strip()

        # Check if title line looks like a variable (no type prefix) — possible cross input
        if not re.match(r'^[skdnmf]\s', title_line):
            # Could be a cross-tab input format or corrupted
            if any(title_line.lower().startswith(v.lower()) for v in df_columns[:50]):
                possible_cross_input = True

        # Check if var line looks like a title (has spaces and no var-like pattern)
        if re.match(r'^[skdnmf]\s', var_line):
            warnings_list.append({
                'level': 'error',
                'msg': f"Red {i+1}: varijabla '{var_line}' izgleda kao naslov tablice. "
                       f"Moguće da je pomaknut red u input skripti.",
                'row': i + 1, 'snippet': _snippet(i),
            })
            continue

        # Extract var names and check existence
        vars_in_line = []
        if var_line.startswith('$'):
            # MR: $e1 '' var1 var2 var3
            parts = var_line.split()
            vars_in_line = [p for p in parts if not p.startswith('$') and p != "''"]
        elif '+' in var_line:
            # Use same half-split logic as actual computation
            vars_in_line = parse_numeric_vars(var_line)
        else:
            vars_in_line = var_line.split() if var_line else []

        for v in vars_in_line:
            if v.lower() not in col_set:
                missing_vars.append((i + 1, v))

        # 3e. Type vs variable-line format mismatch (checked early to gate later checks)
        table_type = title_line[0].lower() if title_line else ''
        is_mr_line = var_line.startswith('$')
        is_numeric_line = '+' in var_line
        is_single_var = not is_mr_line and not is_numeric_line and len(var_line.split()) == 1
        is_multi_var = not is_mr_line and not is_numeric_line and len(var_line.split()) > 1
        has_type_mismatch = False

        if table_type in ('k', 'd') and not is_mr_line:
            warnings_list.append({
                'level': 'warning',
                'msg': f"Red {i+1}: tip '{table_type}' očekuje MR varijable ($e1 ...)",
                'row': i + 1, 'snippet': _snippet(i),
            })
            has_type_mismatch = True
        elif table_type in ('n', 'm') and not is_numeric_line and not is_multi_var:
            warnings_list.append({
                'level': 'warning',
                'msg': f"Red {i+1}: tip '{table_type}' očekuje numeričke varijable (var1 var2 ... var1+var2+...)",
                'row': i + 1, 'snippet': _snippet(i),
            })
            has_type_mismatch = True
        elif table_type == 's' and (is_mr_line or is_numeric_line or is_multi_var):
            warnings_list.append({
                'level': 'warning',
                'msg': f"Red {i+1}: tip 's' očekuje jednu varijablu",
                'row': i + 1, 'snippet': _snippet(i),
            })
            has_type_mismatch = True

        # n/m with "var var" pattern (single-var numeric table) — valid syntax, skip dup/group
        is_nm_single_var = (
            table_type in ('n', 'm') and not is_numeric_line
            and len(vars_in_line) == 2
            and vars_in_line[0].lower() == vars_in_line[1].lower()
        )

        # 3b. Duplicate variables in a line (skip when type mismatch or n/m single-var)
        if not has_type_mismatch and not is_nm_single_var:
            seen_vars = []
            for v in vars_in_line:
                vl = v.lower()
                if vl in seen_vars:
                    warnings_list.append({
                        'level': 'warning',
                        'msg': f"Red {i+1}: varijabla '{v}' se ponavlja u definiciji",
                        'row': i + 1, 'snippet': _snippet(i),
                    })
                else:
                    seen_vars.append(vl)

        # 3c. Title base vs variable base mismatch
        title_base_match = re.match(r'^[skdnmf]\s+([a-zA-Z]+\d+)', title_line)
        if title_base_match and vars_in_line:
            title_base = title_base_match.group(1).lower()
            # Extract bases from variable names (q1_1 → q1, r2_3 → r2)
            var_bases = set()
            base_pat = re.compile(r'^([a-zA-Z]+\d+)[r]?_\d+', re.I)
            for v in vars_in_line:
                bm = base_pat.match(v)
                if bm:
                    var_bases.add(bm.group(1).lower())
            # Also for single vars (like 'r1' without underscore)
            if not var_bases:
                for v in vars_in_line:
                    bm = re.match(r'^([a-zA-Z]+\d+)$', v)
                    if bm:
                        var_bases.add(bm.group(1).lower())
            if var_bases and title_base not in var_bases:
                warnings_list.append({
                    'level': 'warning',
                    'msg': f"Red {i+1}: naslov referira '{title_base}' ali varijable su iz grupe "
                           f"{', '.join(sorted(var_bases))}",
                    'row': i + 1, 'snippet': _snippet(i),
                })

        # 3d. Sum expression consistency check for numeric lines
        if '+' in var_line and not var_line.startswith('$'):
            # Use same half-split logic as actual computation
            _mid = len(var_line) // 2
            _right = var_line[_mid:].strip()
            sum_vars_set = {v.lower() for v in re.split(r'[+\s]+', _right) if v}
            listed_vars = [v.lower() for v in parse_numeric_vars(var_line)]
            listed_set = set(listed_vars)
            not_in_sum = [v for v in listed_vars if v not in sum_vars_set]
            in_sum_not_listed = sorted(v for v in sum_vars_set if v not in listed_set)
            if not_in_sum:
                warnings_list.append({
                    'level': 'warning',
                    'msg': f"Red {i+1}: varijable {', '.join(not_in_sum)} su navedene ali nisu u sumacijskom izrazu",
                    'row': i + 1, 'snippet': _snippet(i),
                })
            if in_sum_not_listed:
                warnings_list.append({
                    'level': 'warning',
                    'msg': f"Red {i+1}: varijable {', '.join(in_sum_not_listed)} su u sumacijskom izrazu ali nisu navedene pojedinačno",
                    'row': i + 1, 'snippet': _snippet(i),
                })

        # 3f. Per-line incomplete group check (skip when type mismatch)
        if not has_type_mismatch and len(vars_in_line) > 1 and _data_base_vars:
            line_bases = {}
            for v in vars_in_line:
                bm = _q_var_pat.match(v)
                if bm:
                    line_bases.setdefault(bm.group(1).lower(), set()).add(v.lower())
            for base, line_vars in line_bases.items():
                if base in _data_base_vars:
                    missing_in_line = sorted(_data_base_vars[base] - line_vars)
                    if missing_in_line:
                        detail = ', '.join(missing_in_line)
                        warnings_list.append({
                            'level': 'warning',
                            'msg': f"Red {i+1}: grupa '{base}' — u datafileu postoje i: {detail}",
                            'row': i + 1, 'snippet': _snippet(i),
                        })

    if possible_cross_input:
        warnings_list.append({
            'level': 'warning',
            'msg': "Neki naslovi izgledaju kao imena varijabli. "
                   "Moguće da je učitan križanje input umjesto total input skripte.",
            'row': None, 'snippet': None,
        })

    # Summarize missing vars
    if missing_vars:
        details = '; '.join(f"red {r}: {v}" for r, v in missing_vars)
        # Build snippets for missing vars
        miss_snippets = []
        for r, v in missing_vars:
            s = _snippet(r - 1)
            if s:
                miss_snippets.append(f"Red {r}:\n{s}")
        warnings_list.append({
            'level': 'warning',
            'msg': f"{len(missing_vars)} varijabli iz input skripte ne postoji u datafileu: {details}",
            'row': missing_vars[0][0] if missing_vars else None,
            'snippet': '\n\n'.join(miss_snippets) if miss_snippets else None,
        })

    # 4. Check if data file has vars that look like they match question patterns in input
    input_var_set = set()
    for var_line in variables:
        var_line = var_line.strip()
        if var_line.startswith('$'):
            parts = var_line.split()
            input_var_set.update(p.lower() for p in parts if not p.startswith('$') and p != "''")
        elif '+' in var_line:
            input_var_set.update(p.lower() for p in var_line.split() if '+' not in p)
        else:
            input_var_set.update(p.lower() for p in var_line.split() if p)

    # Find question-like vars in datafile not covered by input
    q_pattern = re.compile(r'^(q\d+|r\d+)_\d+$', re.I)
    data_q_bases = set()
    input_q_bases = set()
    for c in df_columns:
        # Skip string/text variables — not relevant for table generation
        if df is not None and c in df.columns and pd.api.types.is_object_dtype(df[c]):
            continue
        m = q_pattern.match(c)
        if m:
            data_q_bases.add(m.group(1).lower())
    for v in input_var_set:
        m = q_pattern.match(v)
        if m:
            input_q_bases.add(m.group(1).lower())

    uncovered = data_q_bases - input_q_bases
    if uncovered:
        uncovered_sorted = sorted(uncovered)
        details = ', '.join(uncovered_sorted)
        warnings_list.append({
            'level': 'info',
            'msg': f"Detektirane varijable u datafileu koje nisu u input skripti: {details}",
            'row': None, 'snippet': None,
        })

    return warnings_list


def validate_datafile(df, meta, input_vars=None):
    """
    Validate datafile metadata (value labels, question labels).
    input_vars: set of lowercase variable names used in the input script.
    Returns list of warning dicts: {'level': 'error'|'warning'|'info', 'msg': str}
    """
    warnings_list = []
    if meta is None or df is None:
        return warnings_list

    val_labels_dict = getattr(meta, 'variable_value_labels', {}) or {}
    labels_dict = getattr(meta, 'column_names_to_labels', {}) or {}
    _input_vars = input_vars or set()

    # 1. Values in data without a label (only for vars in input script that have value labels)
    # Group input vars by base (e.g., q22_1, q22_2 → q22) to merge labels like table generation does
    _base_pat = re.compile(r'^([a-zA-Z]+\d+)_\d+$', re.I)
    _base_groups = {}  # base → [var_name, ...]
    col_lower_map = {c.lower(): c for c in df.columns}
    for v_lower in _input_vars:
        actual = col_lower_map.get(v_lower)
        if not actual:
            continue
        bm = _base_pat.match(actual)
        if bm:
            _base_groups.setdefault(bm.group(1).lower(), []).append(actual)

    # Build merged label dicts per base group
    _merged_labels_cache = {}  # base → merged dict
    for base, var_list in _base_groups.items():
        merged = {}
        for vname in var_list:
            vl = val_labels_dict.get(vname, {})
            merged.update(vl)
        _merged_labels_cache[base] = merged

    _checked_bases = set()
    for var_name, vl_dict in val_labels_dict.items():
        if var_name not in df.columns:
            continue
        if _input_vars and var_name.lower() not in _input_vars:
            continue
        # Use merged labels for the base group (mirrors table generation)
        bm = _base_pat.match(var_name)
        if bm:
            base = bm.group(1).lower()
            if base in _checked_bases:
                continue  # already reported for this group
            _checked_bases.add(base)
            check_labels = _merged_labels_cache.get(base, vl_dict)
            # Check all values across all vars in the group
            all_data_vals = set()
            for gvar in _base_groups.get(base, [var_name]):
                if gvar in df.columns:
                    all_data_vals.update(df[gvar].dropna().unique())
        else:
            check_labels = vl_dict
            all_data_vals = set(df[var_name].dropna().unique())

        labeled_keys = set()
        for k in check_labels.keys():
            labeled_keys.add(k)
            try:
                labeled_keys.add(float(k))
            except (ValueError, TypeError):
                pass
            try:
                labeled_keys.add(int(float(k)))
            except (ValueError, TypeError):
                pass
        missing_vals = []
        for v in all_data_vals:
            if v not in labeled_keys:
                missing_vals.append(v)
        if missing_vals:
            missing_vals.sort(key=lambda x: (0, float(x)) if isinstance(x, (int, float)) else (1, str(x)))
            def _fmt_val(v):
                if isinstance(v, float) and v == int(v):
                    return str(int(v))
                return str(v)
            n_missing = len(missing_vals)
            label_src = f"grupa {base}" if bm else var_name
            if n_missing == 1:
                detail = f"vrijednost {_fmt_val(missing_vals[0])} nema labelu"
            elif n_missing <= 4:
                detail = f"vrijednosti {', '.join(_fmt_val(v) for v in missing_vals)} nemaju labele"
            else:
                detail = (f"{n_missing} vrijednosti nemaju labele "
                          f"({', '.join(_fmt_val(v) for v in missing_vals)} ...)")  
            warnings_list.append({
                'level': 'warning',
                'msg': f"{label_src}: {detail}",
                'row': None, 'snippet': None,
            })

    # 1b. Empty variables — all values are NaN (won't appear in banner, filters, etc.)
    for v_lower in _input_vars:
        actual = col_lower_map.get(v_lower)
        if not actual or actual not in df.columns:
            continue
        if df[actual].dropna().empty:
            lbl = labels_dict.get(actual, '')
            disp = f"{actual} ({lbl})" if lbl and lbl != actual else actual
            warnings_list.append({
                'level': 'warning',
                'msg': f"{disp}: varijabla je potpuno prazna (svi redovi su NaN) — neće se pojaviti u banner/filter opcijama",
                'row': None, 'snippet': None,
            })

    # 2. Inconsistent value labels within a variable group
    for base, var_list in _base_groups.items():
        if len(var_list) < 2:
            continue
        # Collect label key sets per variable
        all_keys = set()
        per_var_keys = {}
        for vname in var_list:
            vl = val_labels_dict.get(vname, {})
            keys = frozenset(vl.keys())
            per_var_keys[vname] = keys
            all_keys.update(keys)
        # Find vars missing some labels
        incomplete = []
        for vname in sorted(var_list):
            missing = all_keys - per_var_keys[vname]
            if missing:
                incomplete.append(vname)
        if incomplete:
            if len(incomplete) <= 4:
                vstr = ', '.join(incomplete)
            else:
                vstr = ', '.join(incomplete[:3]) + f' ... (+{len(incomplete)-3})'
            warnings_list.append({
                'level': 'warning',
                'msg': f"grupa {base}: nekonzistentne value labele — {vstr} nemaju sve labele koje imaju ostale varijable u grupi",
                'row': None, 'snippet': None,
            })

    # 3. Suspicious questionnaire phrasing — only flag sub-questions / open-ended prompts
    _suspicious_patterns = [
        # Croatian — "something else, what?" variants
        r'ne\u0161to\s+drugo[,:]?\s*\u0161to',
        r'neka\s+druga[,:]?\s*koja',
        r'neki\s+drugi[,:]?\s*koji',
        r'neko\s+drugo[,:]?\s*koje',
        r'ne\u0161to\s+tre\u0107e[,:]?\s*\u0161to',
        r'\u0161to\s*[?]',
        r'koja\s*[?]',
        r'koji\s*[?]',
        r'koje\s*[?]',
        # English — "other, specify" variants
        r'something\s+else',
        r'other[,:]?\s*(?:please\s+)?specify',
        r'other[,:]?\s*what',
        r'other[,:]?\s*which',
        r'please\s+specify',
    ]
    _suspicious_re = re.compile('|'.join(_suspicious_patterns), re.I)

    for var_name, vl_dict in val_labels_dict.items():
        for val, label in vl_dict.items():
            if _suspicious_re.search(str(label)):
                warnings_list.append({
                    'level': 'warning',
                    'msg': f'{var_name}: value label "{label}"',
                    'row': None, 'snippet': None,
                })
    for var_name, qlabel in labels_dict.items():
        if _suspicious_re.search(str(qlabel)):
            warnings_list.append({
                'level': 'warning',
                'msg': f'{var_name}: question label "{str(qlabel)[:100]}"',
                'row': None, 'snippet': None,
            })

    return warnings_list


# ═══════════════════════════════════════════════════════════════════
#  GRUPIRANJE VARIJABLI IZ INPUT DEFINICIJA
# ═══════════════════════════════════════════════════════════════════

def _extract_group_key(title):
    """Izvuci ključ grupe iz naslova tablice (npr. 'q1' iz 'q1.2. Niže su...')."""
    title = title.strip()
    # Pokušaj matchati pitanje s brojem: q1, q2, r1, q22, dwork, deduc, ...
    m = re.match(r'^([a-zA-Z]+\d+)', title)
    if m:
        return m.group(1).lower()
    # Demografske varijable (Spol:, Dob:, Regija:, ...) → grupa "demo"
    if ':' in title.split('.')[0]:
        return '_demo_'
    # Fallback: prva riječ
    return title.split('.')[0].split(':')[0].split()[0].lower() if title else '_ostalo_'


def _extract_vars_from_line(var_line):
    """Izvuci pojedinačne varijable iz var definicijskog reda."""
    var_line = var_line.strip()
    # MR: $e1 '' var1 var2 var3
    if var_line.startswith('$'):
        return [p for p in var_line.split() if not p.startswith('$') and p != "''"]
    # Numeric composite: q1_1 q1_2 ... q1_1+q1_2+...
    if '+' in var_line:
        return [p for p in var_line.split() if '+' not in p]
    # Simple single var
    return [var_line] if var_line else []


def build_variable_groups(titles, variables, df_columns):
    """
    Gradi hijerarhijsku strukturu varijabli iz input definicija.
    Vraća: OrderedDict { group_key: { 'label': str, 'vars': [str], 'types': set } }
    """
    from collections import OrderedDict
    groups = OrderedDict()
    col_set = {c.lower(): c for c in df_columns}

    for title_line, var_line in zip(titles, variables):
        table_title = get_table_title(title_line)
        table_type = get_table_type(title_line)
        group_key = _extract_group_key(table_title)

        actual_vars = _extract_vars_from_line(var_line)

        if group_key not in groups:
            # Kreiraj label iz prvog naslova u grupi
            if group_key == '_demo_':
                label = '📋 Demografija'
            else:
                # Skrati na zajednički dio pitanja
                base = table_title.split(' - ')[0].strip()
                # Ako ima sub-item (q1.1.), uzmi dio do prvog sub-broja
                m = re.match(r'^([a-zA-Z]+\d+)\.\d+\.\s*(.+)', base)
                if m:
                    label = f"{m.group(1)}. {m.group(2)[:55]}"
                else:
                    label = base[:65]
            groups[group_key] = {'label': label, 'vars': [], 'types': set()}

        groups[group_key]['types'].add(table_type)

        for v in actual_vars:
            resolved = col_set.get(v.lower())
            if resolved and resolved not in groups[group_key]['vars']:
                groups[group_key]['vars'].append(resolved)

    # Dodaj "Ostale" grupu za varijable koje nisu u input.txt
    used = set()
    for g in groups.values():
        used.update(g['vars'])
    other = [c for c in df_columns if c not in used]
    if other:
        groups['_ostalo_'] = {'label': '📦 Ostale varijable', 'vars': other, 'types': set()}

    # Makni prazne grupe
    groups = OrderedDict((k, v) for k, v in groups.items() if v['vars'])

    return groups


# ═══════════════════════════════════════════════════════════════════
#  GENERIRANJE TABLICA (engine wrapper)
# ═══════════════════════════════════════════════════════════════════

def generate_tables(df, meta, titles, variables, weight_col, start_num):
    """Generira sve tablice. Vraća (tables_list, errors_list)."""
    col_map = build_column_map(df)
    count = min(len(titles), len(variables))
    tables = []
    errors = []

    for i in range(count):
        title_line = titles[i]
        var_line = variables[i]
        table_type = get_table_type(title_line)
        table_title = get_table_title(title_line)
        table_num = i + start_num
        title_str = f"{table_title} (Table {table_num}.1)"

        try:
            if table_type == 's':
                result = make_simple_table(df, var_line.strip(), meta, col_map, weight_col)
            elif table_type in ('k', 'd'):
                result = make_mr_table(df, var_line, meta, col_map, table_type, weight_col)
            elif table_type == 'n':
                result = make_numeric_table(df, var_line, meta, col_map, full_stats=True, weight_col=weight_col)
            elif table_type == 'm':
                result = make_numeric_table(df, var_line, meta, col_map, full_stats=False, weight_col=weight_col)
            elif table_type == 'f':
                result = make_freq_table(df, var_line, meta, col_map, weight_col)
            else:
                errors.append(f"Tablica {table_num}: nepoznat tip '{table_type}'")
                continue

            result['title'] = title_str
            result['_idx'] = i
            tables.append(result)
        except KeyError as e:
            errors.append(f"Tablica {table_num}: varijabla {e} ne postoji u podatcima")
        except Exception as e:
            errors.append(f"Tablica {table_num}: {e}")

    return tables, errors


# ═══════════════════════════════════════════════════════════════════
#  FILTER LOGIC
# ═══════════════════════════════════════════════════════════════════

def apply_filter_groups(df, filter_groups):
    """
    Primijeni filter grupe na DataFrame.
    Svaka grupa ima:
      - 'mode': 'single' (jedna varijabla, filtrira po vrijednostima)
                ili 'multi' (grupa varijabli, filtrira po odabranim sub-varijablama)
      - 'logic': 'AND' ili 'OR' (veznik s prethodnom grupom)
    """
    if not filter_groups:
        return df

    try:
        combined_mask = None

        for grp in filter_groups:
            vals = grp.get('vals', [])
            if not vals:
                continue

            mode = grp.get('mode', 'single')
            logic = grp.get('logic', 'AND')

            if mode == 'multi':
                # vals su kodovi; vars su stupci u grupi
                # ispitanik prolazi ako ima odabranu vrijednost u bilo kojem stupcu
                grp_vars = grp.get('vars', [])
                cmp_vals = []
                for v in vals:
                    try:
                        cmp_vals.append(float(v))
                    except (ValueError, TypeError):
                        cmp_vals.append(v)
                grp_mask = pd.Series(False, index=df.index)
                for sv in grp_vars:
                    if sv in df.columns:
                        grp_mask = grp_mask | df[sv].isin(cmp_vals)
            else:
                # single: vals su vrijednosti jedne varijable
                var = grp['var']
                cmp_vals = []
                for v in vals:
                    try:
                        cmp_vals.append(float(v))
                    except (ValueError, TypeError):
                        cmp_vals.append(v)
                grp_mask = df[var].isin(cmp_vals)

            if combined_mask is None:
                combined_mask = grp_mask
            elif logic == 'OR':
                combined_mask = combined_mask | grp_mask
            else:
                combined_mask = combined_mask & grp_mask

        if combined_mask is None:
            return df
        return df[combined_mask].copy()
    except Exception:
        return df


def build_filter_groups_description(filter_groups, labels_dict, val_labels_dict):
    """Napravi citljiv opis filter grupa."""
    parts = []
    for i, grp in enumerate(filter_groups):
        vals = grp.get('vals', [])
        if not vals:
            continue

        mode = grp.get('mode', 'single')
        logic = grp.get('logic', 'AND')
        group_label = grp.get('group_label', '')

        if mode == 'multi':
            val_parts = [labels_dict.get(sv, sv)[:40] for sv in vals]
            val_str = ' ILI '.join(val_parts)
            part = f"`{group_label[:35]}` = {val_str}"
        else:
            var = grp['var']
            var_lbl = labels_dict.get(var) or var
            short_var = var_lbl[:35] if var_lbl != var else var

            vlabels = val_labels_dict.get(var, {})
            val_parts = []
            for v in vals:
                lbl = vlabels.get(v, '')
                if not lbl:
                    try:
                        lbl = vlabels.get(float(v), '')
                    except (ValueError, TypeError):
                        pass
                if not lbl:
                    try:
                        lbl = vlabels.get(int(float(v)), '')
                    except (ValueError, TypeError):
                        pass
                val_parts.append(str(lbl) if lbl else str(v))

            val_str = ' ILI '.join(val_parts)
            part = f"`{short_var}` = {val_str}"

        if i > 0 and parts:
            connector = ' **ILI** ' if logic == 'OR' else ' **I** '
            parts.append(connector)
        parts.append(part)

    return ''.join(parts) if parts else ''


# ═══════════════════════════════════════════════════════════════════
#  PLAN OBRADE : save / load
# ═══════════════════════════════════════════════════════════════════

class _NumpyEncoder(json.JSONEncoder):
    """Handle numpy types when serializing to JSON."""
    def default(self, obj):
        if isinstance(obj, (np.integer,)):
            return int(obj)
        if isinstance(obj, (np.floating,)):
            return float(obj)
        if isinstance(obj, np.ndarray):
            return obj.tolist()
        return super().default(obj)


def collect_plan(output_defs, use_weight, weight_col, start_num,
                 global_filter_groups=None):
    """Build a JSON-serializable dict from the current configuration."""
    plan = {
        'version': 1,
        'global': {
            'use_weight': bool(use_weight),
            'weight_col': weight_col,
            'start_num': int(start_num),
            'add_toc': bool(st.session_state.get('add_toc', False)),
            'filter_groups': global_filter_groups or [],
        },
        'outputs': [],
    }
    for oi, od in enumerate(output_defs):
        out = {
            'type': od['type'],
            'sheet_name': od['sheet_name'],
            'filter_groups': od.get('filter_groups', []),
        }
        out['table_indices'] = od.get('table_indices', [])
        out['table_mode'] = st.session_state.get(f'out_tblmode_{oi}', 'all')
        if od['type'] == 'krizanje':
            out['banner_vars'] = od.get('banner_vars', [])
            out['show_sig'] = od.get('show_sig', False)
            out['show_sig_total'] = od.get('show_sig_total', False)
        plan['outputs'].append(out)
    return plan


def _build_synthetic_entry(tbl, col_lc):
    """
    Create a synthetic (title, variable) pair for an SPO table whose
    variables exist in the dataset but not in input.txt.
    """
    if tbl['is_mr']:
        resolved = [col_lc[v.lower()] for v in tbl['mr_vars'] if v.lower() in col_lc]
        if not resolved:
            return None
        title = f"k [SPO] MR: {resolved[0].split('_')[0]}"
        var_line = "$e1 '' " + ' '.join(resolved)
    elif tbl['is_numeric']:
        resolved = [col_lc[v.lower()] for v in tbl['row_vars'] if v.lower() in col_lc]
        if not resolved:
            return None
        title = f"n [SPO] {resolved[0]}"
        left = ' '.join(resolved)
        right = '+'.join(resolved)
        # parse_numeric_vars splits at len//2 — pad left so midpoint
        # falls on the space between left and right halves
        target_len = len(right)
        left_padded = left.ljust(target_len)
        var_line = left_padded + ' ' + right
    else:
        resolved = [col_lc[v.lower()] for v in tbl['row_vars'] if v.lower() in col_lc]
        if not resolved:
            return None
        title = f"s [SPO] {resolved[0]}"
        var_line = resolved[0]
    return title, var_line


def _build_plan_from_spo(spo_results, titles, variables, df, cat_var_names,
                         all_tbl_indices, loaded_sav_name):
    """
    Convert multiple SPO parse results into a plan dict compatible
    with _apply_plan_outputs().

    Returns (plan_dict, status_list, synthetic_entries) where:
      - status_list has one entry per SPO file
      - synthetic_entries is a list of (title, var_line) tuples to append
    """
    status_list = []
    all_outputs = []
    col_lc = {c.lower(): c for c in df.columns}
    # Work with mutable copies so synthetic entries get proper indices
    titles = list(titles)
    variables = list(variables)
    synthetic_entries = []

    # Build MR group lookup: var_name_lower → list of resolved group vars
    from spss_tables import get_table_type as _gtt, parse_mr_vars as _pmr
    mr_lookup = {}
    for _t, _v in zip(titles, variables):
        if _gtt(_t) in ('k', 'd'):
            _vs = _v.strip()
            raw = _pmr(_vs) if _vs.startswith('$') else _vs.split()
            resolved = [col_lc[x.lower()] for x in raw if x.lower() in col_lc]
            if len(resolved) > 1:
                for rv in resolved:
                    mr_lookup[rv.lower()] = resolved

    for spo in spo_results:
        fname = spo['filename']

        # ── Check SAV compatibility ──
        if not sav_names_match(spo.get('sav_name'), loaded_sav_name):
            status_list.append({
                'filename': fname,
                'ok': False,
                'reason': f"SAV ne odgovara: SPO koristi '{spo.get('sav_name', '?')}', "
                          f"učitan je '{loaded_sav_name}'",
                'n_outputs': 0,
            })
            continue

        if not spo['tables']:
            status_list.append({
                'filename': fname,
                'ok': False,
                'reason': 'Nije pronađena nijedna tablica u SPO datoteci',
                'n_outputs': 0,
            })
            continue

        # ── Match tables to input.txt ──
        matches = match_spo_to_input(spo, titles, variables, df.columns)
        n_ok = 0

        for mi in matches:
            tbl = mi['spo_table']
            if mi['match_status'] == 'no_match':
                continue

            # ── df_only: vars in dataset but not in input.txt → add synthetic entries ──
            if mi['match_status'] == 'df_only':
                if tbl['is_mr']:
                    # One synthetic MR entry for the whole group
                    entry = _build_synthetic_entry(tbl, col_lc)
                    if entry:
                        new_idx = len(titles)
                        titles.append(entry[0])
                        variables.append(entry[1])
                        synthetic_entries.append(entry)
                        mi['matched_indices'] = [new_idx]
                        # Update MR lookup with the new group
                        resolved_mr = [col_lc[v.lower()] for v in tbl['mr_vars']
                                       if v.lower() in col_lc]
                        if len(resolved_mr) > 1:
                            for rv in resolved_mr:
                                mr_lookup[rv.lower()] = resolved_mr
                    else:
                        continue
                elif tbl['is_numeric']:
                    # One synthetic numeric entry
                    entry = _build_synthetic_entry(tbl, col_lc)
                    if entry:
                        new_idx = len(titles)
                        titles.append(entry[0])
                        variables.append(entry[1])
                        synthetic_entries.append(entry)
                        mi['matched_indices'] = [new_idx]
                    else:
                        continue
                else:
                    # One simple entry per row_var
                    new_indices = []
                    for rv in tbl['row_vars']:
                        resolved = col_lc.get(rv.lower())
                        if resolved:
                            new_idx = len(titles)
                            title = f"s [SPO] {resolved}"
                            titles.append(title)
                            variables.append(resolved)
                            synthetic_entries.append((title, resolved))
                            new_indices.append(new_idx)
                    if new_indices:
                        mi['matched_indices'] = new_indices
                    else:
                        continue
                mi['_was_df_only'] = True

            # Build output definition
            out = {
                'type': tbl['type'] or 'total',
                'sheet_name': '',
                'filter_groups': [],
                'table_indices': mi['matched_indices'],
                'table_mode': 'select',
            }

            # ── Filter groups (from SPO filter expression) ──
            filter_expr = tbl.get('filter_expr') or (
                spo['filters'][0] if len(spo['filters']) == 1 else None
            )
            if filter_expr:
                parsed_filters = parse_filter_expression(filter_expr)
                for pf in parsed_filters:
                    var_name = pf['var']
                    resolved = col_lc.get(var_name.lower())
                    if not resolved:
                        continue
                    vals_f = [float(v) if isinstance(v, int) else v for v in pf['vals']]
                    # Check if var is part of an MR group → multi filter
                    mr_group = mr_lookup.get(resolved.lower())
                    if mr_group:
                        out['filter_groups'].append({
                            'mode': 'multi',
                            'vars': mr_group,
                            'vals': vals_f,
                            'logic': 'AND',
                        })
                    else:
                        out['filter_groups'].append({
                            'mode': 'single',
                            'var': resolved,
                            'vals': vals_f,
                            'logic': 'AND',
                        })

            # ── Banner (for krizanje) ──
            if tbl['type'] == 'krizanje' and tbl.get('banner_var'):
                bvar = col_lc.get(tbl['banner_var'].lower())
                if bvar and bvar in cat_var_names:
                    out['banner_vars'] = [bvar]
                    out['show_sig'] = True
                    out['show_sig_total'] = False
                else:
                    # Banner var not categorical — fall back to total
                    out['type'] = 'total'

            # ── Sheet name ──
            base = 'TOTAL' if out['type'] == 'total' else 'CROSS'
            if out.get('banner_vars'):
                base = 'CROSS_' + '_'.join(out['banner_vars'])
            if out['filter_groups']:
                fg0 = out['filter_groups'][0]
                if fg0.get('mode') == 'multi':
                    # Use first var of the MR group for naming
                    fvar = fg0.get('vars', [''])[0]
                else:
                    fvar = fg0.get('var', '')
                fvals = fg0.get('vals', [])
                fstr = f"{fvar}={'_'.join(str(int(v)) for v in fvals)}"
                base = f"{base}_{fstr}"
            out['sheet_name'] = base[:31]

            all_outputs.append(out)
            n_ok += 1

        if n_ok > 0:
            n_synth = sum(1 for m in matches
                          if m.get('_was_df_only'))
            msg = f'{n_ok} output(a) prepoznato'
            if n_synth:
                msg += f' ({n_synth} dodano u input skriptu)'
            status_list.append({
                'filename': fname,
                'ok': True,
                'reason': msg,
                'n_outputs': n_ok,
            })
        else:
            reasons = set(m['reason'] for m in matches
                          if m['match_status'] in ('no_match', 'df_only'))
            reason_str = '; '.join(reasons) if reasons else 'Varijable iz SPO ne odgovaraju input.txt tablicama'
            status_list.append({
                'filename': fname,
                'ok': False,
                'reason': reason_str[:200],
                'n_outputs': 0,
            })

    plan = {
        'version': 1,
        'global': {
            'use_weight': False,
            'weight_col': '',
            'start_num': 1,
            'add_toc': False,
            'filter_groups': [],
        },
        'outputs': all_outputs,
    }
    return plan, status_list, synthetic_entries


def _apply_plan_outputs(plan, cat_var_names, filter_choices,
                        all_tbl_indices, df, val_labels_dict):
    """Set session_state keys for outputs from a loaded plan."""
    outputs = plan.get('outputs', [])
    st.session_state['n_outputs'] = max(len(outputs), 1)

    for oi, out in enumerate(outputs):
        st.session_state[f'out_type_{oi}'] = out.get('type', 'total')
        st.session_state[f'out_name_{oi}'] = out.get('sheet_name', f'Output_{oi+1}')
        st.session_state[f'out_name_dirty_{oi}'] = True

        # ── Filters ──
        fg = out.get('filter_groups', [])
        st.session_state[f'out_filt_{oi}'] = bool(fg)
        if fg:
            st.session_state[f'n_fg_{oi}'] = len(fg)
            for fi, fgroup in enumerate(fg):
                if fi > 0:
                    logic_val = "ILI (OR)" if fgroup.get('logic') == 'OR' else "I (AND)"
                    st.session_state[f'fg_logic_{oi}_{fi}'] = logic_val

                if fgroup.get('mode') == 'multi':
                    saved_vars = set(fgroup.get('vars', []))
                    fc_idx = None
                    for j, fc in enumerate(filter_choices):
                        if fc['mode'] == 'multi' and set(fc['vars']) == saved_vars:
                            fc_idx = j
                            break
                    if fc_idx is None:
                        continue  # skip filter group — vars no longer exist
                    st.session_state[f'fg_var_{oi}_{fi}'] = fc_idx
                    matched_fc = filter_choices[fc_idx]
                    # Build _all_vals the same way the UI does
                    _all_vals_set = set()
                    for _sv in matched_fc['vars']:
                        if _sv in df.columns:
                            for _uv in df[_sv].dropna().unique():
                                _all_vals_set.add(_uv)
                    try:
                        _all_vals = sorted(
                            _all_vals_set,
                            key=lambda x: (0, float(x)) if isinstance(x, (int, float)) else (1, str(x))
                        )
                    except TypeError:
                        _all_vals = sorted(_all_vals_set, key=str)
                    val_idx = []
                    for sv in fgroup.get('vals', []):
                        for k, uv in enumerate(_all_vals):
                            if uv == sv or str(uv) == str(sv):
                                val_idx.append(k)
                                break
                    st.session_state[f'fg_vals_{oi}_{fi}'] = val_idx
                else:
                    the_var = fgroup.get('var', '')
                    fc_idx = None
                    for j, fc in enumerate(filter_choices):
                        if fc['mode'] == 'single' and fc['vars'][0] == the_var:
                            fc_idx = j
                            break
                    if fc_idx is None:
                        continue  # skip filter group — var no longer exists
                    st.session_state[f'fg_var_{oi}_{fi}'] = fc_idx
                    # Map saved values to indices in unique_vals
                    try:
                        unique_vals = sorted(
                            df[the_var].dropna().unique(),
                            key=lambda x: (0, float(x)) if isinstance(x, (int, float)) else (1, str(x))
                        )
                    except (TypeError, KeyError):
                        unique_vals = []
                    val_idx = []
                    for sv in fgroup.get('vals', []):
                        for k, uv in enumerate(unique_vals):
                            if uv == sv or str(uv) == str(sv):
                                val_idx.append(k)
                                break
                    st.session_state[f'fg_vals_{oi}_{fi}'] = val_idx

        # ── Table selection (both types) ──
        tbl_mode = out.get('table_mode', 'all')
        st.session_state[f'out_tblmode_{oi}'] = tbl_mode
        saved_set = set(out.get('table_indices', []))
        if tbl_mode == 'exclude':
            st.session_state[f'out_excl_{oi}'] = [
                j for j, idx in enumerate(all_tbl_indices) if idx not in saved_set]
        elif tbl_mode == 'select':
            st.session_state[f'out_sel_{oi}'] = [
                j for j, idx in enumerate(all_tbl_indices) if idx in saved_set]

        # ── Križanje settings ──
        if out.get('type') == 'krizanje':
            banner_idx = [cat_var_names.index(v)
                          for v in out.get('banner_vars', []) if v in cat_var_names]
            st.session_state[f'out_banner_{oi}'] = banner_idx
            st.session_state[f'out_sig_{oi}'] = out.get('show_sig', True)
            st.session_state[f'out_sigtot_{oi}'] = out.get('show_sig_total', False)


# ═══════════════════════════════════════════════════════════════════
#  STREAMLIT UI
# ═══════════════════════════════════════════════════════════════════

def _auto_sheet_name(out_type, banner_indices, cat_var_names, use_weight):
    """Generate automatic sheet name from output type and banner selection."""
    if out_type == 'total':
        base = 'TOTAL'
    elif not banner_indices:
        base = 'CROSS'
    else:
        parts = []
        for bi in banner_indices:
            if bi < len(cat_var_names):
                parts.append(cat_var_names[bi].lower())
        base = ('CROSS_' + '_'.join(parts)) if parts else 'CROSS'
    return base


def main():
    _assets = os.path.join(os.path.dirname(__file__), 'assets')
    _favicon_path = os.path.join(_assets, 'favicon.png')
    st.set_page_config(
        page_title="Hendalice",
        page_icon=_favicon_path if os.path.exists(_favicon_path) else "📊",
        layout="wide",
        initial_sidebar_state="collapsed",
    )

    # ══════════════════════════════════════════════════
    #  SIDEBAR: Info
    # ══════════════════════════════════════════════════
    with st.sidebar:
        st.markdown("### Hendalice")
        st.markdown("""
        **Koraci:**

        1. Učitajte **.sav** i **input.txt**
        2. Po potrebi uključite **ponder**
        3. Kreirajte **outpute** (Total / Križanje)
        4. **Generirajte** Excel

        ---

        **💾 Plan obrade:**
        Spremite konfiguraciju kao _po.json
        i kasnije je učitajte za nastavak.

        ---

        **Tipovi tablica:**
        | Tip | Opis |
        |-----|------|
        | s | Frekvencija (n, %) |
        | k | Multiple Response |
        | d | Multi Dichotomy |
        | n | Numerička (full) |
        | m | Numerička (mean) |
        | f | Frequencies |
        """)

        st.divider()
        st.caption("Hendalice v2.4")

    # ── Custom CSS ──
    st.markdown("""
    <style>
    .header-row {
        display: flex;
        align-items: center;
        gap: 0.7rem;
        margin-bottom: 0.2rem;
    }
    .header-row h1 {
        margin: 0;
        font-size: 2rem;
        font-weight: 700;
    }
    </style>
    """, unsafe_allow_html=True)

    # ── Header ──
    _logo_path = os.path.join(_assets, 'logo.svg')
    if os.path.exists(_logo_path):
        with open(_logo_path, 'r', encoding='utf-8') as _f:
            _logo_svg = _f.read()
    else:
        _logo_svg = ''
    st.markdown(
        '<div class="header-row">'
        f'{_logo_svg}'
        '<h1>Hendalice</h1>'
        '</div>',
        unsafe_allow_html=True,
    )

    st.divider()

    # ══════════════════════════════════════════════════
    #  KORAK 1: Upload fajlova
    # ══════════════════════════════════════════════════

    # ── Widget-key preservation across file swaps ──
    # Streamlit GC-s widget-backed session_state keys when the widget isn't
    # rendered.  When the user clears a file uploader to replace it, the
    # early-return path prevents Korak 2/3 widgets from rendering, losing
    # all output settings.  We snapshot them before returning early and
    # restore them once both files are loaded again.
    _WIDGET_PREFIXES = (
        'out_type_', 'out_name_', 'out_filt_', 'out_banner_',
        'out_sig_', 'out_sigtot_', 'out_tblmode_', 'out_excl_', 'out_sel_',
        'out_name_dirty_', 'out_autoname_',
        'fg_logic_', 'fg_var_', 'fg_vals_', 'n_fg_',
        'gfg_logic_', 'gfg_var_', 'gfg_vals_',
    )
    _WIDGET_GLOBALS = ('use_weight', 'weight_idx', 'add_toc', 'table_design',
                       'n_outputs', '_out_order', 'global_filt', 'n_gfg')

    def _snapshot_widget_config():
        # Don't overwrite an existing snapshot (e.g. saved by _reset_data
        # before Streamlit GC'd the widget keys on the next rerun)
        if '_widget_snap' in st.session_state:
            return
        snap = {}
        for k in list(st.session_state.keys()):
            if k.startswith(_WIDGET_PREFIXES) or k in _WIDGET_GLOBALS:  # type: ignore[union-attr]
                snap[k] = st.session_state[k]
        if snap:
            st.session_state['_widget_snap'] = snap

    def _restore_widget_config():
        snap = st.session_state.pop('_widget_snap', None)
        if snap:
            for k, v in snap.items():
                if k not in st.session_state:
                    st.session_state[k] = v

    def _invalidate_data_indices():
        """Clear index-based session state keys that depend on data/input file contents."""
        # Use max of current and any previous n_outputs to clear all remnants
        _n_out = max(st.session_state.get('n_outputs', 1),
                     st.session_state.get('_prev_n_outputs', 1))
        for _oi in range(_n_out):
            for _k in (f'out_banner_{_oi}', f'out_excl_{_oi}', f'out_sel_{_oi}'):
                st.session_state.pop(_k, None)
            for _fi in range(st.session_state.get(f'n_fg_{_oi}', 0) + 1):
                st.session_state.pop(f'fg_var_{_oi}_{_fi}', None)
                st.session_state.pop(f'fg_vals_{_oi}_{_fi}', None)
                st.session_state.pop(f'fg_logic_{_oi}_{_fi}', None)
        st.session_state.pop('weight_idx', None)
        # Clear global filter indices
        for _fi in range(st.session_state.get('n_gfg', 0) + 1):
            st.session_state.pop(f'gfg_var_{_fi}', None)
            st.session_state.pop(f'gfg_vals_{_fi}', None)
            st.session_state.pop(f'gfg_logic_{_fi}', None)
        # Remember peak n_outputs for future invalidation
        st.session_state['_prev_n_outputs'] = _n_out
        # Also clean from snapshot
        snap = st.session_state.get('_widget_snap')
        if snap:
            for _k in list(snap.keys()):
                if _k.startswith(('out_banner_', 'out_excl_', 'out_sel_',
                                  'fg_var_', 'fg_vals_', 'fg_logic_',
                                  'gfg_var_', 'gfg_vals_', 'gfg_logic_')) or _k == 'weight_idx':
                    del snap[_k]

    def _reset_data():
        """Clear loaded data files from session, keep outputs & plan."""
        _snapshot_widget_config()
        for k in list(st.session_state.keys()):
            if k.startswith(('df', 'meta', '_sav', 'titles', 'variables',  # type: ignore[union-attr]
                             'break_vars', '_input', 'var_groups')):
                del st.session_state[k]
        # Rotate only the data-uploader key (sav + input.txt)
        st.session_state['_data_ugen'] = st.session_state.get('_data_ugen', 0) + 1

    def _reset_settings():
        """Clear plan, outputs & settings — keep data files."""
        n_prev = st.session_state.get('n_outputs', 1)
        # Remove plan
        for k in list(st.session_state.keys()):
            if k.startswith(('_plan_applied', '_pending_plan')):  # type: ignore[union-attr]
                del st.session_state[k]
        # Rotate plan uploader key
        st.session_state['_plan_ugen'] = st.session_state.get('_plan_ugen', 0) + 1
        # Clear SPO import state
        st.session_state.pop('_spo_applied_batch', None)
        st.session_state.pop('_spo_status', None)
        st.session_state['_spo_ugen'] = st.session_state.get('_spo_ugen', 0) + 1
        # Reset global settings
        st.session_state['use_weight'] = False
        st.session_state['add_toc'] = False
        st.session_state['table_design'] = 'hendal'
        # Reset global filter
        st.session_state['global_filt'] = False
        n_gfg_prev = st.session_state.get('n_gfg', 0)
        st.session_state.pop('n_gfg', None)
        for _fi in range(n_gfg_prev + 1):
            st.session_state.pop(f'gfg_var_{_fi}', None)
            st.session_state.pop(f'gfg_vals_{_fi}', None)
            st.session_state.pop(f'gfg_logic_{_fi}', None)
        # Reset outputs
        st.session_state['n_outputs'] = 1
        st.session_state['_out_order'] = [0]
        for oi in range(max(n_prev, 1)):
            st.session_state[f'out_type_{oi}'] = 'total'
            st.session_state[f'out_filt_{oi}'] = False
            st.session_state.pop(f'out_sig_{oi}', None)
            st.session_state[f'out_sigtot_{oi}'] = False
            st.session_state[f'out_banner_{oi}'] = []
            st.session_state[f'out_name_dirty_{oi}'] = False
            for k in (f'out_name_{oi}', f'out_autoname_{oi}',
                      f'out_tblmode_{oi}', f'out_excl_{oi}', f'out_sel_{oi}',
                      f'n_fg_{oi}'):
                st.session_state.pop(k, None)
        st.session_state.pop('_widget_snap', None)

    def _reset_all():
        """Clear everything — data, outputs, all settings."""
        dgen = st.session_state.get('_data_ugen', 0) + 1
        pgen = st.session_state.get('_plan_ugen', 0) + 1
        spogen = st.session_state.get('_spo_ugen', 0) + 1
        n_prev = st.session_state.get('n_outputs', 1)
        for k in list(st.session_state.keys()):
            del st.session_state[k]
        st.session_state['_data_ugen'] = dgen
        st.session_state['_plan_ugen'] = pgen
        st.session_state['_spo_ugen'] = spogen
        # Explicitly set widget keys to defaults so Streamlit's internal
        # widget cache is overridden on next render
        st.session_state['use_weight'] = False
        st.session_state['add_toc'] = False
        st.session_state['table_design'] = 'hendal'
        st.session_state['global_filt'] = False
        st.session_state['n_outputs'] = 1
        st.session_state['_out_order'] = [0]
        # Reset output widgets for all previously active outputs
        for oi in range(max(n_prev, 1)):
            st.session_state[f'out_type_{oi}'] = 'total'
            st.session_state[f'out_filt_{oi}'] = False
            st.session_state.pop(f'out_sig_{oi}', None)
            st.session_state[f'out_sigtot_{oi}'] = False
            st.session_state[f'out_banner_{oi}'] = []
            st.session_state[f'out_name_dirty_{oi}'] = False
            for k in (f'out_name_{oi}', f'out_autoname_{oi}',
                      f'out_tblmode_{oi}', f'out_excl_{oi}', f'out_sel_{oi}',
                      f'n_fg_{oi}'):
                st.session_state.pop(k, None)

    hdr1, _, btn_rd, btn_rs, btn_ra = st.columns([6, 0.5, 1.2, 1.2, 1.2])
    with hdr1:
        st.header("1. Učitajte podatke")
    with btn_rd:
        st.markdown("<div style='height:0.8rem'></div>", unsafe_allow_html=True)
        st.button("🔄 Reset podataka", key="_btn_reset_data",
                  on_click=_reset_data, help="Makni učitane fajlove, zadrži plan i outpute")
    with btn_rs:
        st.markdown("<div style='height:0.8rem'></div>", unsafe_allow_html=True)
        st.button("🔄 Reset postavki", key="_btn_reset_settings",
                  on_click=_reset_settings, help="Makni plan, outpute i postavke, zadrži podatke")
    with btn_ra:
        st.markdown("<div style='height:0.8rem'></div>", unsafe_allow_html=True)
        st.button("🗑️ Reset svega", key="_btn_reset_all",
                  on_click=_reset_all, help="Makni sve — podatke, outpute, postavke")

    col_sav, col_input = st.columns(2)
    _dugen = st.session_state.get('_data_ugen', 0)

    with col_sav:
        st.subheader("📁 SPSS podatkovni fajl (.sav)")
        sav_file = st.file_uploader(
            "Odaberite .sav fajl",
            type=["sav"],
            help="SPSS podatkovni fajl s vašim podacima",
            key=f"sav_upload_{_dugen}",
        )

    with col_input:
        st.subheader("📝 Definicije tablica (input.txt)")
        input_file = st.file_uploader(
            "Odaberite input.txt",
            type=["txt"],
            help="Tekstualni fajl s definicijama tablica (3 sekcije odvojene praznim redom)",
            key=f"input_upload_{_dugen}",
        )

    # ── Učitaj .sav ako je uploadano ──
    if sav_file is not None:
        sav_bytes = sav_file.read()
        sav_hash = hashlib.md5(sav_bytes).hexdigest()
        if 'df' not in st.session_state or st.session_state.get('_sav_hash') != sav_hash:
            with st.spinner("Učitavam .sav podatke..."):
                with tempfile.NamedTemporaryFile(suffix='.sav', delete=False) as tmp:
                    tmp.write(sav_bytes)
                    tmp_path = tmp.name
                try:
                    df, meta = pyreadstat.read_sav(tmp_path, apply_value_formats=False)
                    st.session_state['df'] = df
                    st.session_state['meta'] = meta
                    st.session_state['_sav_name'] = sav_file.name
                    st.session_state['_sav_hash'] = sav_hash
                finally:
                    os.unlink(tmp_path)
            _invalidate_data_indices()

        df = st.session_state['df']
        meta = st.session_state['meta']

        st.success(f"✅ Učitano: **{sav_file.name}** — {len(df)} ispitanika, {len(df.columns)} varijabli")
    else:
        df = None
        meta = None

    # ── Učitaj input.txt ako je uploadano ──
    if input_file is not None:
        raw = input_file.read()
        input_hash = hashlib.md5(raw).hexdigest()
        if 'titles' not in st.session_state or st.session_state.get('_input_hash') != input_hash:
            break_vars, titles, variables = parse_input_bytes(raw)
            st.session_state['break_vars'] = break_vars
            st.session_state['titles'] = titles
            st.session_state['variables'] = variables
            st.session_state['_input_name'] = input_file.name
            st.session_state['_input_hash'] = input_hash
            _invalidate_data_indices()

        titles = st.session_state['titles']
        variables = st.session_state['variables']
        st.success(f"✅ Učitano: **{input_file.name}** — {len(titles)} tablica definirano")
    else:
        titles = None
        variables = None

    # ── Plan obrade (opcionalno) ──
    st.subheader("📋 Plan obrade")
    _pugen = st.session_state.get('_plan_ugen', 0)
    plan_file = st.file_uploader(
        "Učitajte prethodno spremljenu konfiguraciju (_po.json)",
        type=["json"],
        help="Prethodno spremljeni plan obrade sa svim postavkama",
        key=f"plan_upload_{_pugen}",
    )
    if plan_file is not None and df is not None and titles is not None:
        if st.session_state.get('_plan_applied_name') != plan_file.name:
            try:
                plan_data = json.loads(plan_file.read().decode('utf-8'))
                st.session_state['_pending_plan'] = plan_data
                st.session_state['_plan_applied_name'] = plan_file.name
            except (json.JSONDecodeError, UnicodeDecodeError):
                st.error("❌ Neispravan JSON fajl.")

    # ── SPO import (opcionalno) ──
    st.subheader("📊 Import iz SPO")
    _spogen = st.session_state.get('_spo_ugen', 0)
    spo_files = st.file_uploader(
        "Učitajte .spo datoteke za automatsko prepoznavanje outputa",
        type=["spo"],
        accept_multiple_files=True,
        help="Uploadajte jedan ili više SPSS Output (.spo) fajlova. "
             "Aplikacija će analizirati svaki i pokušati rekonstruirati postavke outputa.",
        key=f"spo_upload_{_spogen}",
    )
    if spo_files and df is not None and titles is not None:
        # Track which batch we already processed
        _spo_batch_key = '|'.join(sorted(f.name for f in spo_files))
        if st.session_state.get('_spo_applied_batch') != _spo_batch_key:
            spo_results = []
            for spo_f in spo_files:
                try:
                    spo_bytes = spo_f.read()
                    tmp_path = os.path.join(tempfile.gettempdir(), spo_f.name)
                    with open(tmp_path, 'wb') as fout:
                        fout.write(spo_bytes)
                    result = parse_spo(tmp_path)
                    spo_results.append(result)
                    os.unlink(tmp_path)
                except Exception as exc:
                    spo_results.append({
                        'filename': spo_f.name,
                        'sav_name': None,
                        'filters': [],
                        'tables': [],
                        '_error': str(exc),
                    })

            loaded_sav_name = st.session_state.get('_sav_name', '')
            plan_from_spo, spo_status, spo_synthetics = _build_plan_from_spo(
                spo_results, titles, variables, df,
                # cat_var_names not ready yet — compute quickly
                [c for c in df.columns if 2 <= df[c].dropna().nunique() <= 30],
                [i for i, t in enumerate(titles)
                 if get_table_type(t) in ('s', 'k', 'd', 'n', 'm')],
                loaded_sav_name,
            )

            # Append synthetic input entries to session_state
            if spo_synthetics:
                ext_titles = list(st.session_state['titles'])
                ext_vars = list(st.session_state['variables'])
                for syn_title, syn_var in spo_synthetics:
                    ext_titles.append(syn_title)
                    ext_vars.append(syn_var)
                st.session_state['titles'] = ext_titles
                st.session_state['variables'] = ext_vars
                titles = ext_titles
                variables = ext_vars

            st.session_state['_spo_status'] = spo_status
            st.session_state['_spo_applied_batch'] = _spo_batch_key

            if plan_from_spo['outputs']:
                st.session_state['_pending_plan'] = plan_from_spo
                st.session_state['_plan_applied_name'] = f'_spo_import_{_spo_batch_key}'

    # Show SPO status (persists across reruns)
    if '_spo_status' in st.session_state:
        for s in st.session_state['_spo_status']:
            if s['ok']:
                st.success(f"✅ **{s['filename']}** — {s['reason']}")
            else:
                st.error(f"❌ **{s['filename']}** — {s['reason']}")

    if df is None or titles is None:
        _snapshot_widget_config()
        st.info("👆 Učitajte oba fajla za nastavak.")
        return

    _restore_widget_config()

    # ── Validacija input skripte ──
    break_vars = st.session_state.get('break_vars', [])
    _validation_warnings = validate_input(titles, variables, list(df.columns), break_vars, df=df, meta=meta)
    if _validation_warnings:
        n_err = sum(1 for w in _validation_warnings if w['level'] == 'error')
        n_warn = sum(1 for w in _validation_warnings if w['level'] == 'warning')
        n_info = sum(1 for w in _validation_warnings if w['level'] == 'info')
        parts = []
        if n_err:
            parts.append(f"{n_err} grešaka")
        if n_warn:
            parts.append(f"{n_warn} upozorenja")
        if n_info:
            parts.append(f"{n_info} info")
        summary = ', '.join(parts)
        with st.expander(f"⚠️ {summary} o input skripti", expanded=True):
            for w in _validation_warnings:
                snippet = w.get('snippet')
                if w['level'] == 'error':
                    st.error(w['msg'])
                elif w['level'] == 'warning':
                    st.warning(w['msg'])
                else:
                    st.info(w['msg'])
                if snippet:
                    st.code(snippet, language=None)

    # ── Validacija datafile-a ──
    _input_var_set = set()
    for var_line in variables:  # type: ignore[union-attr]
        var_line_s = var_line.strip()
        if var_line_s.startswith('$'):
            _input_var_set.update(p.lower() for p in var_line_s.split() if not p.startswith('$') and p != "''")
        elif '+' in var_line_s:
            _input_var_set.update(p.lower() for p in var_line_s.split() if '+' not in p)
        else:
            _input_var_set.update(p.lower() for p in var_line_s.split() if p)
    _df_warnings = validate_datafile(df, meta, input_vars=_input_var_set)
    if _df_warnings:
        n_warn_d = len(_df_warnings)
        with st.expander(f"⚠️ {n_warn_d} upozorenja o datafileu", expanded=False):
            for w in _df_warnings:
                st.warning(w['msg'])

    st.divider()

    # ══════════════════════════════════════════════════
    #  KORAK 2: Postavke
    # ══════════════════════════════════════════════════
    st.header("2. Postavke")

    all_vars = list(df.columns)
    labels_dict = getattr(meta, 'column_names_to_labels', {}) or {}
    val_labels_dict = getattr(meta, 'variable_value_labels', {}) or {}
    numeric_vars = [c for c in all_vars if pd.api.types.is_numeric_dtype(df[c])]

    def var_display(v):
        lbl = labels_dict.get(v, '')
        if lbl and lbl != v:
            return f"{v}  —  {lbl[:60]}"
        return v

    # ── Pre-apply global settings from pending plan ──
    _pp = st.session_state.get('_pending_plan')
    if _pp:
        _g = _pp.get('global', {})
        st.session_state['use_weight'] = _g.get('use_weight', False)
        st.session_state['add_toc'] = _g.get('add_toc', False)
        wc = _g.get('weight_col')
        if wc and wc in numeric_vars:
            st.session_state['weight_idx'] = numeric_vars.index(wc)

    st.subheader("⚖️ Ponder")

    # Auto-detect pond variable
    _pond_candidates = [c for c in all_vars if c.lower() in ('pond', 'ponder', 'weight')]
    if _pond_candidates and not st.session_state.get('use_weight', False):
        pond_var = _pond_candidates[0]
        st.info(f"💡 Detektirana ponder varijabla **{pond_var}** u datafileu, ali ponder nije uključen.")

    use_weight = st.checkbox("Koristi ponder", key="use_weight")
    weight_col = None
    if use_weight:
        # Pre-select pond variable if detected
        default_idx = 0
        if _pond_candidates:
            for pi, nv in enumerate(numeric_vars):
                if nv == _pond_candidates[0]:
                    default_idx = pi
                    break
            if 'weight_idx' not in st.session_state:
                st.session_state['weight_idx'] = default_idx

        weight_idx = st.selectbox(
            "Odaberite varijablu pondera:",
            options=range(len(numeric_vars)),
            format_func=lambda i: var_display(numeric_vars[i]),
            key="weight_idx",
        )
        weight_col = numeric_vars[weight_idx]
        st.caption(f"Suma: {df[weight_col].sum():.1f} | Prosjek: {df[weight_col].mean():.4f}")

    start_num = 1

    st.subheader("📑 Table of Contents")
    add_toc = st.checkbox("Dodaj TOC sheet (BETA)", key="add_toc",
                          help="Dodaje početni sheet s popisom svih tablica i linkovima")

    st.subheader("🎨 Dizajn tablica")
    table_design = st.selectbox(
        "Stil:",
        options=['hendal', 'mate'],
        format_func=lambda x: {'hendal': '🟡 Hendal', 'mate': '🐉 Mate'}[x],
        key="table_design",
    )

    # ── Priprema za filtre i krizanja ──
    var_groups = build_variable_groups(titles, variables, df.columns)
    st.session_state['var_groups'] = var_groups

    # Kategoricke varijable za banner
    cat_vars = []
    for v in df.columns:
        nunique = df[v].dropna().nunique()
        if 2 <= nunique <= 30:
            lbl = labels_dict.get(v) or ''
            disp = f"{v} — {lbl}" if lbl and lbl != v else v
            cat_vars.append((v, disp))
    cat_var_names = [cv[0] for cv in cat_vars]
    cat_var_displays = [cv[1] for cv in cat_vars]

    # Tablice iz input.txt
    table_options = []
    for i, t in enumerate(titles):
        tt = get_table_type(t)
        if tt in ('s', 'k', 'd', 'n', 'm'):
            tn = get_table_title(t)
            table_options.append((i, f"T{i + start_num} [{tt}] {tn[:70]}"))
    all_tbl_indices = [to[0] for to in table_options]
    all_tbl_displays = [to[1] for to in table_options]

    # Filter choices — built from input script titles for accurate labels
    filter_choices = []
    _filt_seen_vars = set()
    col_set_lc = {c.lower(): c for c in df.columns}
    for i, (title_line, var_line) in enumerate(zip(titles, variables)):  # type: ignore[arg-type]
        ttype = get_table_type(title_line)
        ttitle = get_table_title(title_line)
        actual_vars = _extract_vars_from_line(var_line)
        resolved = [col_set_lc[v.lower()] for v in actual_vars if v.lower() in col_set_lc]
        if not resolved:
            continue
        # Multi-response (k/d) → one multi-select entry
        if ttype in ('k', 'd') and len(resolved) > 1:
            key = frozenset(resolved)
            if key not in _filt_seen_vars:
                _filt_seen_vars.add(key)
                filter_choices.append({
                    'display': f"T{i+start_num} [{ttype}] {ttitle}",
                    'mode': 'multi',
                    'vars': resolved,
                })
        else:
            # Single var — use input script title
            v = resolved[0]
            if v not in _filt_seen_vars:
                _filt_seen_vars.add(v)
                filter_choices.append({
                    'display': f"T{i+start_num} [{ttype}] {ttitle}",
                    'mode': 'single',
                    'vars': [v],
                })
    # Ostale varijable iz datafile-a koje nisu u inputu
    _used_vars = set()
    for fc in filter_choices:
        _used_vars.update(fc['vars'])
    for v in df.columns:
        if v not in _used_vars:
            nunique = df[v].dropna().nunique()
            if 2 <= nunique <= 30:
                lbl = labels_dict.get(v) or ''
                disp = f"{v} — {lbl}" if lbl and lbl != v else v
                filter_choices.append({
                    'display': f"[df] {disp}",
                    'mode': 'single',
                    'vars': [v],
                })
    choice_displays = [fc['display'] for fc in filter_choices]

    # ── Apply pending plan (global filter + output settings) ──
    # Must run BEFORE global filter UI to avoid modifying instantiated widgets
    if '_pending_plan' in st.session_state:
        _pp_data = st.session_state.pop('_pending_plan')
        # Restore global filter from plan
        _gfg = _pp_data.get('global', {}).get('filter_groups', [])
        st.session_state['global_filt'] = bool(_gfg)
        if _gfg:
            st.session_state['n_gfg'] = len(_gfg)
            for _fi, _fgroup in enumerate(_gfg):
                if _fi > 0:
                    _logic_val = "ILI (OR)" if _fgroup.get('logic') == 'OR' else "I (AND)"
                    st.session_state[f'gfg_logic_{_fi}'] = _logic_val
                if _fgroup.get('mode') == 'multi':
                    _saved_vars = set(_fgroup.get('vars', []))
                    _fc_idx = None
                    for _j, _fc in enumerate(filter_choices):
                        if _fc['mode'] == 'multi' and set(_fc['vars']) == _saved_vars:
                            _fc_idx = _j
                            break
                    if _fc_idx is None:
                        continue  # skip — vars no longer exist
                    st.session_state[f'gfg_var_{_fi}'] = _fc_idx
                    _matched_fc = filter_choices[_fc_idx]
                    # Build _all_vals the same way the UI does
                    _all_vals_set = set()
                    for _sv in _matched_fc['vars']:
                        if _sv in df.columns:
                            for _uv in df[_sv].dropna().unique():
                                _all_vals_set.add(_uv)
                    try:
                        _all_vals = sorted(
                            _all_vals_set,
                            key=lambda x: (0, float(x)) if isinstance(x, (int, float)) else (1, str(x))
                        )
                    except TypeError:
                        _all_vals = sorted(_all_vals_set, key=str)
                    _val_idx = []
                    for _sv in _fgroup.get('vals', []):
                        for _k, _uv in enumerate(_all_vals):
                            if _uv == _sv or str(_uv) == str(_sv):
                                _val_idx.append(_k)
                                break
                    st.session_state[f'gfg_vals_{_fi}'] = _val_idx
                else:
                    _the_var = _fgroup.get('var', '')
                    _fc_idx = None
                    for _j, _fc in enumerate(filter_choices):
                        if _fc['mode'] == 'single' and _fc['vars'][0] == _the_var:
                            _fc_idx = _j
                            break
                    if _fc_idx is None:
                        continue  # skip — var no longer exists
                    st.session_state[f'gfg_var_{_fi}'] = _fc_idx
                    try:
                        _unique_vals = sorted(
                            df[_the_var].dropna().unique(),
                            key=lambda x: (0, float(x)) if isinstance(x, (int, float)) else (1, str(x))
                        )
                    except (TypeError, KeyError):
                        _unique_vals = []
                    _val_idx = []
                    for _sv in _fgroup.get('vals', []):
                        for _k, _uv in enumerate(_unique_vals):
                            if _uv == _sv or str(_uv) == str(_sv):
                                _val_idx.append(_k)
                                break
                    st.session_state[f'gfg_vals_{_fi}'] = _val_idx
        _apply_plan_outputs(
            _pp_data,
            cat_var_names, filter_choices, all_tbl_indices,
            df, val_labels_dict,
        )
        st.rerun()

    # ── Globalni filter ──
    st.subheader("🔍 Globalni filter")
    st.caption("Primjenjuje se na **sve** outpute. Kombinira se s per-output filterima koristeći AND.")
    global_filt = st.checkbox("Koristi globalni filter", key="global_filt")
    global_filter_groups = []
    if global_filt and filter_choices:
        if 'n_gfg' not in st.session_state:
            st.session_state['n_gfg'] = 1

        def _add_gfg():
            st.session_state['n_gfg'] += 1
        def _rm_gfg():
            if st.session_state['n_gfg'] > 1:
                st.session_state['n_gfg'] -= 1

        for gfi in range(st.session_state['n_gfg']):
            if gfi > 0:
                g_logic = st.radio(
                    "Veznik", ["I (AND)", "ILI (OR)"],
                    key=f"gfg_logic_{gfi}",
                    horizontal=True,
                    label_visibility="collapsed",
                )
                g_row_logic = 'OR' if 'ILI' in g_logic else 'AND'
            else:
                g_row_logic = 'AND'

            gc_var, gc_vals = st.columns([2, 3])

            with gc_var:
                def _on_gfg_var_change(_fi=gfi):
                    vals_key = f'gfg_vals_{_fi}'
                    if vals_key in st.session_state:
                        st.session_state[vals_key] = []

                g_choice_idx = st.selectbox(
                    "Pitanje",
                    options=range(len(filter_choices)),
                    format_func=lambda i, cd=choice_displays: cd[i],
                    key=f"gfg_var_{gfi}",
                    label_visibility="collapsed",
                    on_change=_on_gfg_var_change,
                )
                g_chosen = filter_choices[g_choice_idx]

            with gc_vals:
                if g_chosen['mode'] == 'multi':
                    g_sub_vars = g_chosen['vars']
                    _g_all_vals_set = set()
                    _g_val_label_map = {}
                    for sv in g_sub_vars:
                        for uv in df[sv].dropna().unique():
                            _g_all_vals_set.add(uv)
                        svl = val_labels_dict.get(sv, {})
                        for code, lbl in svl.items():
                            if lbl:
                                _g_val_label_map[code] = lbl
                    try:
                        _g_all_vals = sorted(
                            _g_all_vals_set,
                            key=lambda x: (0, float(x)) if isinstance(x, (int, float)) else (1, str(x))
                        )
                    except TypeError:
                        _g_all_vals = sorted(_g_all_vals_set, key=str)

                    def _g_mr_val_display(uv, vlm=_g_val_label_map):
                        lbl = vlm.get(uv, '')
                        if not lbl:
                            try:
                                lbl = vlm.get(int(uv) if isinstance(uv, float) and uv == int(uv) else uv, '')
                            except (ValueError, TypeError):
                                pass
                        return f"{uv} — {lbl}" if lbl else str(uv)

                    _g_mr_displays = [_g_mr_val_display(v) for v in _g_all_vals]

                    g_selected = st.multiselect(
                        "Vrijednosti",
                        options=range(len(_g_all_vals)),
                        format_func=lambda i, md=_g_mr_displays: md[i],
                        key=f"gfg_vals_{gfi}",
                        label_visibility="collapsed",
                        placeholder="Odaberite...",
                    )
                    g_selected_vals = [_g_all_vals[i] for i in g_selected]
                    global_filter_groups.append({
                        'mode': 'multi',
                        'group_label': g_chosen['display'],
                        'vars': g_sub_vars,
                        'vals': g_selected_vals,
                        'logic': g_row_logic,
                    })
                else:
                    g_the_var = g_chosen['vars'][0]
                    g_var_vlabels = val_labels_dict.get(g_the_var, {})
                    try:
                        g_unique_vals = sorted(
                            df[g_the_var].dropna().unique(),
                            key=lambda x: (0, float(x)) if isinstance(x, (int, float)) else (1, str(x))
                        )
                    except TypeError:
                        g_unique_vals = sorted(df[g_the_var].dropna().unique(), key=str)

                    g_val_options = []
                    for uv in g_unique_vals:
                        lbl = g_var_vlabels.get(uv, '')
                        if not lbl:
                            try:
                                lbl = g_var_vlabels.get(int(uv) if isinstance(uv, float) and uv == int(uv) else uv, '')
                            except (ValueError, TypeError):
                                pass
                        g_val_options.append(f"{uv} — {lbl}" if lbl else str(uv))

                    g_selected = st.multiselect(
                        "Vrijednosti",
                        options=range(len(g_unique_vals)),
                        format_func=lambda i, vo=g_val_options: vo[i],
                        key=f"gfg_vals_{gfi}",
                        label_visibility="collapsed",
                        placeholder="Odaberite...",
                    )
                    g_selected_vals = [g_unique_vals[i] for i in g_selected]
                    global_filter_groups.append({
                        'mode': 'single',
                        'var': g_the_var,
                        'group_label': g_chosen['display'],
                        'vals': g_selected_vals,
                        'logic': g_row_logic,
                    })

        gbc1, gbc2, _ = st.columns([1, 1, 4])
        with gbc1:
            st.button("➕ Uvjet", on_click=_add_gfg, key="gfg_add")
        with gbc2:
            st.button("➖ Ukloni", on_click=_rm_gfg, key="gfg_rm")

        active_gfg = [g for g in global_filter_groups if g.get('vals')]
        if active_gfg:
            try:
                gfilt_df = apply_filter_groups(df, active_gfg)
                st.caption(f"Globalni filter: **{len(gfilt_df)}** od {len(df)} ispitanika")
            except Exception:
                pass

    # Spremi aktivne globalne grupe za korištenje u generiranju
    active_global_filter_groups = [g for g in global_filter_groups if g.get('vals')]

    st.divider()

    # ══════════════════════════════════════════════════
    #  KORAK 3: Outputi
    # ══════════════════════════════════════════════════

    def _reset_outputs():
        """Reset all outputs back to a single Total."""
        n = st.session_state.get('n_outputs', 1)
        for oi in range(n):
            for key_tpl in ('out_type_{}', 'out_name_{}', 'out_filt_{}',
                            'out_banner_{}', 'out_sig_{}', 'out_sigtot_{}',
                            'out_tblmode_{}', 'out_excl_{}', 'out_sel_{}',
                            'n_fg_{}', 'out_name_dirty_{}', 'out_autoname_{}'):
                k = key_tpl.format(oi)
                if k in st.session_state:
                    del st.session_state[k]
            # Also clean filter group keys
            for fi in range(st.session_state.get(f'n_fg_{oi}', 0) + 1):
                for fg_tpl in ('fg_logic_{}_{}', 'fg_var_{}_{}', 'fg_vals_{}_{}'):
                    k = fg_tpl.format(oi, fi)
                    if k in st.session_state:
                        del st.session_state[k]
        st.session_state['n_outputs'] = 1
        st.session_state['_out_order'] = [0]

    hdr3, _, btn_rst = st.columns([6, 2, 1.2])
    with hdr3:
        st.header("3. Outputi")
    with btn_rst:
        st.markdown("<div style='height:0.8rem'></div>", unsafe_allow_html=True)
        st.button("🗑️ Reset outputa", key="_btn_reset_outputs",
                  on_click=_reset_outputs, help="Obriši sve outpute i vrati na jedan Total")
    st.caption("Svaki output postaje Excel sheet. Može biti **Total** ili **Križanje** (banner tablice).")

    if 'n_outputs' not in st.session_state:
        st.session_state['n_outputs'] = 1

    def _add_output():
        n = st.session_state['n_outputs']
        st.session_state['n_outputs'] = n + 1
        # Add new item to order
        order = st.session_state.get('_out_order', list(range(n)))
        order.append(n)
        st.session_state['_out_order'] = order

    def _delete_output(target_oi):
        """Delete a specific output and compact remaining indices."""
        n = st.session_state['n_outputs']
        if n <= 1:
            return
        order = st.session_state.get('_out_order', list(range(n)))
        if target_oi not in order:
            return
        order = [x for x in order if x != target_oi]

        # Build remap: sorted remaining old indices → new contiguous indices
        remaining_sorted = sorted(set(order))
        remap = {old: new for new, old in enumerate(remaining_sorted)}

        # Record max n_fg for each output before clearing
        max_fgs = {oi: st.session_state.get(f'n_fg_{oi}', 0) for oi in range(n)}

        # Collect state for remaining outputs
        saved = {}
        for old_idx in remaining_sorted:
            new_idx = remap[old_idx]
            for key_tpl in ('out_type_{}', 'out_name_{}', 'out_filt_{}',
                            'out_banner_{}', 'out_sig_{}', 'out_sigtot_{}',
                            'out_tblmode_{}', 'out_excl_{}', 'out_sel_{}',
                            'n_fg_{}', 'out_name_dirty_{}', 'out_autoname_{}'):
                k = key_tpl.format(old_idx)
                if k in st.session_state:
                    saved[key_tpl.format(new_idx)] = st.session_state[k]
            for fi in range(max_fgs[old_idx] + 1):
                for fg_tpl in ('fg_logic_{}_{}', 'fg_var_{}_{}', 'fg_vals_{}_{}'):
                    k = fg_tpl.format(old_idx, fi)
                    if k in st.session_state:
                        saved[fg_tpl.format(new_idx, fi)] = st.session_state[k]

        # Clear all output keys
        for oi in range(n):
            for key_tpl in ('out_type_{}', 'out_name_{}', 'out_filt_{}',
                            'out_banner_{}', 'out_sig_{}', 'out_sigtot_{}',
                            'out_tblmode_{}', 'out_excl_{}', 'out_sel_{}',
                            'n_fg_{}', 'out_name_dirty_{}', 'out_autoname_{}'):
                st.session_state.pop(key_tpl.format(oi), None)
            for fi in range(max_fgs.get(oi, 0) + 1):
                for fg_tpl in ('fg_logic_{}_{}', 'fg_var_{}_{}', 'fg_vals_{}_{}'):
                    st.session_state.pop(fg_tpl.format(oi, fi), None)

        # Restore saved state with new indices
        for k, v in saved.items():
            st.session_state[k] = v

        st.session_state['_out_order'] = [remap[x] for x in order]
        st.session_state['n_outputs'] = n - 1

    def _remove_output():
        n = st.session_state['n_outputs']
        if n > 1:
            order = st.session_state.get('_out_order', list(range(n)))
            _delete_output(order[-1])

    def _duplicate_output(src):
        """Copy all session_state keys from output *src* to a new output after src."""
        n = st.session_state['n_outputs']
        st.session_state['n_outputs'] = n + 1
        dst = n  # 0-based index of the new output

        # Keys that are directly per-output
        for key_tpl in ('out_type_{}', 'out_name_{}', 'out_filt_{}',
                        'out_banner_{}', 'out_sig_{}', 'out_sigtot_{}',
                        'out_tblmode_{}', 'out_excl_{}', 'out_sel_{}',
                        'n_fg_{}'):
            k_src = key_tpl.format(src)
            if k_src in st.session_state:
                st.session_state[key_tpl.format(dst)] = st.session_state[k_src]

        # Append suffix to sheet name to avoid duplicate
        name_key = f'out_name_{dst}'
        if name_key in st.session_state:
            st.session_state[name_key] = st.session_state[name_key] + '_kopija'
        st.session_state[f'out_name_dirty_{dst}'] = True

        # Filter group rows
        n_fg = st.session_state.get(f'n_fg_{src}', 0)
        for fi in range(n_fg):
            for fg_tpl in ('fg_logic_{}_{}', 'fg_var_{}_{}', 'fg_vals_{}_{}'):
                k_src = fg_tpl.format(src, fi)
                if k_src in st.session_state:
                    st.session_state[fg_tpl.format(dst, fi)] = st.session_state[k_src]

        # Insert duplicate right after source in display order
        order = st.session_state.get('_out_order', list(range(n)))
        src_pos = order.index(src) if src in order else len(order)
        order.insert(src_pos + 1, dst)
        st.session_state['_out_order'] = order

    output_defs = []

    n_out_total = st.session_state['n_outputs']

    # ── Arrow-based reordering ──
    if '_out_order' not in st.session_state:
        st.session_state['_out_order'] = list(range(n_out_total))
    order = st.session_state['_out_order']
    # Ensure order is consistent with n_outputs
    existing = set(range(n_out_total))
    order = [x for x in order if x in existing]
    for x in existing - set(order):
        order.append(x)
    st.session_state['_out_order'] = order

    def _swap_outputs(pos_a, pos_b):
        """Swap two outputs by their position in the order list."""
        o = st.session_state['_out_order']
        o[pos_a], o[pos_b] = o[pos_b], o[pos_a]
        st.session_state['_out_order'] = o

    for pos, oi in enumerate(order):
        with st.container(border=True):
            # ── Header: Output N ──
            if n_out_total > 1:
                up_col, dn_col, h_col, type_col, name_col, reset_col, dup_col, del_col = st.columns([0.18, 0.18, 0.7, 2, 2, 0.35, 0.35, 0.35])
                with up_col:
                    st.button("▲", key=f"up_{oi}", disabled=(pos == 0),
                              on_click=_swap_outputs, args=(pos, pos - 1),
                              help="Pomakni gore")
                with dn_col:
                    st.button("▼", key=f"dn_{oi}", disabled=(pos == n_out_total - 1),
                              on_click=_swap_outputs, args=(pos, pos + 1),
                              help="Pomakni dolje")
            else:
                h_col, type_col, name_col, reset_col, dup_col = st.columns([0.8, 2, 2, 0.35, 0.35])

            with h_col:
                st.markdown(f"### Output {pos + 1}")

            with dup_col:
                st.button("📋", key=f"dup_{oi}",
                          on_click=_duplicate_output, args=(oi,),
                          help="Dupliciraj ovaj output")

            if n_out_total > 1:
                with del_col:
                    st.button("🗑️", key=f"del_{oi}",
                              on_click=_delete_output, args=(oi,),
                              help="Obriši ovaj output")

            with type_col:
                out_type = st.selectbox(
                    "Tip:",
                    options=['total', 'krizanje'],
                    format_func=lambda t: {'total': '📋 Total (frekvencije)',
                                           'krizanje': '📊 Križanje (banner)'}[t],
                    key=f"out_type_{oi}",
                )

            # ── Auto-name logic ──
            prev_banner = st.session_state.get(f'out_banner_{oi}', [])
            auto_name_full = _auto_sheet_name(out_type, prev_banner, cat_var_names, use_weight)
            auto_name = auto_name_full[:31]
            is_dirty = st.session_state.get(f'out_name_dirty_{oi}', False)
            prev_auto = st.session_state.get(f'out_autoname_{oi}', '')

            if not is_dirty and (f'out_name_{oi}' not in st.session_state or prev_auto != auto_name):
                st.session_state[f'out_name_{oi}'] = auto_name
            st.session_state[f'out_autoname_{oi}'] = auto_name

            if len(auto_name_full) > 31 and not is_dirty:
                st.caption(f"ℹ️ Auto ime odrezano na 31 znak: **{auto_name}** (puno ime: {auto_name_full})")

            def _on_name_change(_oi=oi):
                cur = st.session_state.get(f'out_name_{_oi}', '')
                auto = st.session_state.get(f'out_autoname_{_oi}', '')
                if cur != auto:
                    st.session_state[f'out_name_dirty_{_oi}'] = True

            def _reset_name(_oi=oi):
                st.session_state[f'out_name_dirty_{_oi}'] = False
                # Force the auto name to regenerate
                if f'out_name_{_oi}' in st.session_state:
                    del st.session_state[f'out_name_{_oi}']

            with name_col:
                sheet_name = st.text_input("Ime sheeta:",
                                           key=f"out_name_{oi}",
                                           on_change=_on_name_change)
                if len(sheet_name) > 31:
                    st.warning(f"⚠️ Ime sheeta ima {len(sheet_name)} znakova — Excel limit je 31. Bit će odrezano na: **{sheet_name[:31]}**")

            with reset_col:
                st.markdown("<div style='height:1.6rem'></div>", unsafe_allow_html=True)
                if is_dirty:
                    st.button("↻", key=f"reset_name_{oi}",
                              on_click=_reset_name,
                              help="Vrati automatski generirano ime")

            # ── Filter (per-output) ──
            with st.expander("🔍 Filter", expanded=False):
                use_filt = st.checkbox("Koristi filter",
                                       key=f"out_filt_{oi}")
                out_filter_groups = []
                if use_filt and filter_choices:
                    if f'n_fg_{oi}' not in st.session_state:
                        st.session_state[f'n_fg_{oi}'] = 1

                    def _add_fg(o=oi):
                        st.session_state[f'n_fg_{o}'] += 1
                    def _rm_fg(o=oi):
                        if st.session_state[f'n_fg_{o}'] > 1:
                            st.session_state[f'n_fg_{o}'] -= 1

                    for fi in range(st.session_state[f'n_fg_{oi}']):
                        if fi > 0:
                            logic = st.radio(
                                "Veznik", ["I (AND)", "ILI (OR)"],
                                key=f"fg_logic_{oi}_{fi}",
                                horizontal=True,
                                label_visibility="collapsed",
                            )
                            row_logic = 'OR' if 'ILI' in logic else 'AND'
                        else:
                            row_logic = 'AND'

                        c_var, c_vals = st.columns([2, 3])

                        with c_var:
                            def _on_fg_var_change(_oi=oi, _fi=fi):
                                vals_key = f'fg_vals_{_oi}_{_fi}'
                                if vals_key in st.session_state:
                                    st.session_state[vals_key] = []

                            choice_idx = st.selectbox(
                                "Pitanje",
                                options=range(len(filter_choices)),
                                format_func=lambda i, cd=choice_displays: cd[i],
                                key=f"fg_var_{oi}_{fi}",
                                label_visibility="collapsed",
                                on_change=_on_fg_var_change,
                            )
                            chosen = filter_choices[choice_idx]

                        with c_vals:
                            if chosen['mode'] == 'multi':
                                # Collect all unique values + labels across all vars in group
                                sub_vars = chosen['vars']
                                _all_vals_set = set()
                                _val_label_map = {}
                                for sv in sub_vars:
                                    for uv in df[sv].dropna().unique():
                                        _all_vals_set.add(uv)
                                    svl = val_labels_dict.get(sv, {})
                                    for code, lbl in svl.items():
                                        if lbl:
                                            _val_label_map[code] = lbl
                                try:
                                    _all_vals = sorted(
                                        _all_vals_set,
                                        key=lambda x: (0, float(x)) if isinstance(x, (int, float)) else (1, str(x))
                                    )
                                except TypeError:
                                    _all_vals = sorted(_all_vals_set, key=str)

                                def _mr_val_display(uv, vlm=_val_label_map):
                                    lbl = vlm.get(uv, '')
                                    if not lbl:
                                        try:
                                            lbl = vlm.get(int(uv) if isinstance(uv, float) and uv == int(uv) else uv, '')
                                        except (ValueError, TypeError):
                                            pass
                                    return f"{uv} — {lbl}" if lbl else str(uv)

                                _mr_displays = [_mr_val_display(v) for v in _all_vals]

                                selected = st.multiselect(
                                    "Vrijednosti",
                                    options=range(len(_all_vals)),
                                    format_func=lambda i, md=_mr_displays: md[i],
                                    key=f"fg_vals_{oi}_{fi}",
                                    label_visibility="collapsed",
                                    placeholder="Odaberite...",
                                )
                                selected_vals = [_all_vals[i] for i in selected]
                                out_filter_groups.append({
                                    'mode': 'multi',
                                    'group_label': chosen['display'],
                                    'vars': sub_vars,
                                    'vals': selected_vals,
                                    'logic': row_logic,
                                })
                            else:
                                the_var = chosen['vars'][0]
                                var_vlabels = val_labels_dict.get(the_var, {})
                                try:
                                    unique_vals = sorted(
                                        df[the_var].dropna().unique(),
                                        key=lambda x: (0, float(x)) if isinstance(x, (int, float)) else (1, str(x))
                                    )
                                except TypeError:
                                    unique_vals = sorted(df[the_var].dropna().unique(), key=str)

                                val_options = []
                                for uv in unique_vals:
                                    lbl = var_vlabels.get(uv, '')
                                    if not lbl:
                                        try:
                                            lbl = var_vlabels.get(int(uv) if isinstance(uv, float) and uv == int(uv) else uv, '')
                                        except (ValueError, TypeError):
                                            pass
                                    val_options.append(f"{uv} — {lbl}" if lbl else str(uv))

                                selected = st.multiselect(
                                    "Vrijednosti",
                                    options=range(len(unique_vals)),
                                    format_func=lambda i, vo=val_options: vo[i],
                                    key=f"fg_vals_{oi}_{fi}",
                                    label_visibility="collapsed",
                                    placeholder="Odaberite...",
                                )
                                selected_vals = [unique_vals[i] for i in selected]
                                out_filter_groups.append({
                                    'mode': 'single',
                                    'var': the_var,
                                    'group_label': chosen['display'],
                                    'vals': selected_vals,
                                    'logic': row_logic,
                                })

                    bc1, bc2, _ = st.columns([1, 1, 4])
                    with bc1:
                        st.button("➕ Uvjet", on_click=_add_fg, key=f"fg_add_{oi}")
                    with bc2:
                        st.button("➖ Ukloni", on_click=_rm_fg, key=f"fg_rm_{oi}")

                    active_fg = [g for g in out_filter_groups if g.get('vals')]
                    if active_fg:
                        try:
                            filt_df = apply_filter_groups(df, active_fg)
                            st.caption(f"Filter: **{len(filt_df)}** od {len(df)} ispitanika")
                        except Exception:
                            pass

            # ── Križanje specifično: banner + tablice ──
            banner_vars_sel = []
            show_sig = False
            show_sig_total = False
            tbl_final_indices = list(all_tbl_indices)

            if out_type == 'krizanje':
                if not cat_vars:
                    st.warning("Nema kategoričkih varijabli za banner.")
                elif not table_options:
                    st.warning("Nema tablica za križanje.")
                else:
                    banner_vars_sel = st.multiselect(
                        "Banner varijable:",
                        options=range(len(cat_vars)),
                        format_func=lambda i, cd=cat_var_displays: cd[i],
                        key=f"out_banner_{oi}",
                        placeholder="Odaberite banner varijable...",
                    )

                    if banner_vars_sel:
                        for bi in banner_vars_sel:
                            bvar = cat_var_names[bi]
                            bvl = val_labels_dict.get(bvar, {})
                            bvals = sorted(
                                df[bvar].dropna().unique(),
                                key=lambda x: (0, float(x)) if isinstance(x, (int, float)) else (1, str(x))
                            )
                            cats = []
                            for bv in bvals[:8]:
                                lbl = bvl.get(bv, '')
                                if not lbl:
                                    try:
                                        lbl = bvl.get(int(bv) if isinstance(bv, float) and bv == int(bv) else bv, '')
                                    except (ValueError, TypeError):
                                        pass
                                cats.append(str(lbl) if lbl else str(bv))
                            more = '...' if len(bvals) > 8 else ''
                            st.caption(f"↳ **{bvar}** ({len(bvals)} kat.): {', '.join(cats)}{more}")

                    st.session_state.setdefault(f'out_sig_{oi}', True)
                    show_sig = st.checkbox("Značajnost (z-test 95%)",
                                           key=f"out_sig_{oi}",
                                           help="Generira dodatni sheet s _sig sufiksom")

                    show_sig_total = st.checkbox("Sig Total (Total vs kategorije)",
                                                key=f"out_sigtot_{oi}",
                                                help="Generira dodatni sheet s _sig_total — Total stupac se testira protiv svake kategorije")

            # ── Odabir tablica (za oba tipa) ──
            if table_options:
                tbl_mode = st.radio(
                    "Tablice:",
                    options=['all', 'exclude', 'select'],
                    format_func=lambda m: {
                        'all': f'Sve ({len(table_options)})',
                        'exclude': 'Sve osim isključenih',
                        'select': 'Samo odabrane',
                    }[m],
                    key=f"out_tblmode_{oi}",
                    horizontal=True,
                )

                if tbl_mode == 'exclude':
                    excluded = st.multiselect(
                        "Isključi:",
                        options=range(len(table_options)),
                        format_func=lambda i, td=all_tbl_displays: td[i],
                        key=f"out_excl_{oi}",
                    )
                    tbl_final_indices = [all_tbl_indices[j] for j in range(len(table_options)) if j not in excluded]
                elif tbl_mode == 'select':
                    selected = st.multiselect(
                        "Odaberi:",
                        options=range(len(table_options)),
                        format_func=lambda i, td=all_tbl_displays: td[i],
                        key=f"out_sel_{oi}",
                    )
                    tbl_final_indices = [all_tbl_indices[j] for j in selected]

            # ── Spremi output definiciju ──
            out_def = {
                'type': out_type,
                'sheet_name': sheet_name[:31],
                'filter_groups': [g for g in out_filter_groups if g.get('vals')] if use_filt else [],
                'table_indices': tbl_final_indices,
            }
            if out_type == 'krizanje':
                out_def['banner_vars'] = [cat_var_names[bi] for bi in banner_vars_sel]
                out_def['show_sig'] = show_sig
                out_def['show_sig_total'] = show_sig_total
            output_defs.append(out_def)

    bc1, bc2, _ = st.columns([1, 1, 4])
    with bc1:
        st.button("➕ Dodaj output", on_click=_add_output)
    with bc2:
        st.button("➖ Ukloni output", on_click=_remove_output)

    # ── Spremi plan obrade ──
    if output_defs:
        plan_dict = collect_plan(output_defs, use_weight, weight_col, start_num,
                                global_filter_groups=active_global_filter_groups)
        plan_json = json.dumps(plan_dict, ensure_ascii=False, indent=2, cls=_NumpyEncoder)
        sav_base = os.path.splitext(st.session_state.get('_sav_name', 'data'))[0]
        plan_filename = f"{sav_base}_po.json"
        st.download_button(
            label="💾 Spremi plan obrade",
            data=plan_json,
            file_name=plan_filename,
            mime="application/json",
        )

    st.divider()

    # ══════════════════════════════════════════════════
    #  KORAK 4: Pregled i generiranje
    # ══════════════════════════════════════════════════
    st.header("4. Generiraj tablice")

    # Pregled input-a
    with st.expander("📋 Pregled definiranih tablica", expanded=False):
        preview_data = []
        type_names = {
            's': 'Frekvencija', 'k': 'Multiple Response', 'd': 'Multi Dichotomy',
            'n': 'Numerička (full)', 'm': 'Numerička (mean)', 'f': 'Frequencies',
        }
        for i, t in enumerate(titles):
            tt = get_table_type(t)
            tn = get_table_title(t)
            preview_data.append({
                '#': i + start_num,
                'Tip': type_names.get(tt, tt),
                'Naslov': tn[:80],
                'Var': variables[i][:50] if i < len(variables) else '',  # type: ignore[arg-type]
            })
        st.dataframe(preview_data, use_container_width=True, hide_index=True)

        # ── Prikaži par pravih tablica ──
        st.markdown("---")
        st.markdown("**Primjeri stvarnih tablica** (prvih 5)")
        col_map = build_column_map(df)
        _shown = 0
        for i in range(min(len(titles), len(variables))):  # type: ignore[arg-type]
            if _shown >= 5:
                break
            title_line = titles[i]
            var_line = variables[i]  # type: ignore[index]
            tt = get_table_type(title_line)
            tn = get_table_title(title_line)
            try:
                if tt == 's':
                    tbl = make_simple_table(df, var_line.strip(), meta, col_map, weight_col)
                elif tt in ('k', 'd'):
                    tbl = make_mr_table(df, var_line, meta, col_map, tt, weight_col)
                elif tt in ('n', 'm'):
                    tbl = make_numeric_table(df, var_line, meta, col_map, tt == 'n', weight_col)
                else:
                    continue
                tbl_df = pd.DataFrame(tbl['rows'], columns=tbl['header'])
                st.caption(f"**T{i + start_num} [{tt}]** {tn[:70]}")
                st.dataframe(tbl_df, use_container_width=True, hide_index=True, height=min(len(tbl_df) * 35 + 40, 300))
                _shown += 1
            except Exception:
                continue

    # Sažetak outputa
    if active_global_filter_groups:
        gf_desc = build_filter_groups_description(active_global_filter_groups, labels_dict, val_labels_dict)
        st.info(f"🔍 **Globalni filter aktivan:** {gf_desc}")
    out_summary = []
    for od in output_defs:
        filt_str = f" + filter ({len(od['filter_groups'])} uvjeta)" if od['filter_groups'] else ""
        if od['type'] == 'total':
            nt = len(od.get('table_indices', []))
            tbl_str = f" ({nt} tablica)" if nt < len(table_options) else ""
            out_summary.append(f"**{od['sheet_name']}** — Total{tbl_str}{filt_str}")
        else:
            nb = len(od.get('banner_vars', []))
            nt = len(od.get('table_indices', []))
            sig_str = (" + _sig" if od.get('show_sig') else "") + \
                      (" + _sig_total" if od.get('show_sig_total') else "")
            out_summary.append(f"**{od['sheet_name']}** — Križanje ({nb} bannera × {nt} tablica){sig_str}{filt_str}")
    st.markdown(" · ".join(out_summary) if out_summary else "Nema definiranih outputa.")

    # ── Gumb za generiranje ──
    if st.button("🚀 Generiraj Excel tablice", type="primary", use_container_width=True):

        progress = st.progress(0, text="Pripremam podatke...")
        t_start = time.time()

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name

        from openpyxl import Workbook as _Workbook
        wb = _Workbook()
        # Ukloni default sheet
        wb.remove(wb.active)  # type: ignore[arg-type]
        existing_sheets = []

        total_tables = 0
        total_xt = 0
        all_errors = []

        # ── Pre-build TOC question structure ──
        toc_rows = []        # ordered list of question dicts
        toc_positions = {}   # table_idx → list of {'sheet', 'row', 'cat', 'kriz_idx'}
        toc_kriz_sheets = [] # list of {'plain', 'sig', 'sigT'} per kriz output
        if add_toc:
            import re as _re
            _base_to_row = {}
            _mean_pending = []

            for _ti in range(len(titles)):
                _ttype = get_table_type(titles[_ti])
                _ttitle = get_table_title(titles[_ti])
                _stripped = _ttitle.rstrip()
                _is_t2b = '- T2B' in _stripped
                _is_mean = _ttype in ('n', 'm')

                if _is_mean:
                    _mv = _re.match(r'^(\S+)', _ttitle.strip())
                    _mean_var = _mv.group(1).rstrip('.:') if _mv else ''
                    _mean_pending.append((_ti, _mean_var))
                elif _is_t2b:
                    _base = _stripped[:_stripped.index('- T2B')].rstrip()  # remove "- T2B" and everything after
                    if _base in _base_to_row:
                        toc_rows[_base_to_row[_base]]['t2b_idx'] = _ti
                    else:
                        _base_to_row[_base] = len(toc_rows)
                        toc_rows.append({'title': _base, 'regular_idx': None,
                                         'mean_idx': None, 't2b_idx': _ti})
                else:
                    _base = _stripped
                    if _base not in _base_to_row:
                        _base_to_row[_base] = len(toc_rows)
                        toc_rows.append({'title': _base, 'regular_idx': _ti,
                                         'mean_idx': None, 't2b_idx': None})

            # Link MEAN tables to sub-questions by prefix
            for _mi, _mvar in _mean_pending:
                _matched = False
                if _mvar:
                    for _row in toc_rows:
                        _rm = _re.match(r'^(\S+)', _row['title'].strip())
                        _rv = _rm.group(1).rstrip('.:') if _rm else ''
                        if _rv and (
                            _rv == _mvar or
                            (_rv.startswith(_mvar) and len(_rv) > len(_mvar) and _rv[len(_mvar)] == '.')
                        ):
                            _row['mean_idx'] = _mi
                            _matched = True
                if not _matched:
                    _mt = get_table_title(titles[_mi]).rstrip()
                    if '- MEAN' in _mt:
                        _mt = _mt[:_mt.index('- MEAN')].rstrip()
                    if _mt not in _base_to_row:
                        _base_to_row[_mt] = len(toc_rows)
                        toc_rows.append({'title': _mt, 'regular_idx': None,
                                         'mean_idx': _mi, 't2b_idx': None})

        n_out = len(output_defs)
        for out_i, out_def in enumerate(output_defs):
            pct_base = int(10 + 80 * out_i / max(n_out, 1))
            progress.progress(pct_base, text=f"Output {out_i + 1}/{n_out}: {out_def['sheet_name']}...")

            # Primijeni globalni + per-output filter
            work_df = df.copy()
            if active_global_filter_groups:
                work_df = apply_filter_groups(work_df, active_global_filter_groups)
            if out_def['filter_groups']:
                work_df = apply_filter_groups(work_df, out_def['filter_groups'])

            # ── Unique sheet name ──
            def _unique_name(base):
                name = base[:31]
                while name in existing_sheets:
                    sfx = 2
                    while f"{name[:28]}_{sfx}" in existing_sheets:
                        sfx += 1
                    name = f"{name[:28]}_{sfx}"
                existing_sheets.append(name)
                return name

            if out_def['type'] == 'total':
                # ── TOTAL output → frekvencijske tablice ──
                tbl_indices_set = set(out_def.get('table_indices', []))
                tables, errs = generate_tables(
                    work_df, meta, titles, variables, weight_col, start_num
                )
                if tbl_indices_set:
                    tables = [t for t in tables if t.get('_idx') in tbl_indices_set]
                total_tables += len(tables)
                all_errors.extend(errs)

                # Piši u privremeni fajl pa kopiraj sheetove
                import tempfile as _tmpmod
                with _tmpmod.NamedTemporaryFile(suffix='.xlsx', delete=False) as _tf:
                    _tf_path = _tf.name
                write_tables_to_excel(tables, _tf_path, design=table_design)
                _twb = load_workbook(_tf_path)
                for _srcws in _twb.worksheets:
                    sname = _unique_name(out_def['sheet_name'])
                    _destws = wb.create_sheet(title=sname)
                    for row in _srcws.iter_rows():
                        for cell in row:
                            _destws.cell(row=cell.row, column=cell.column, value=cell.value)
                            if cell.has_style:
                                _destws.cell(row=cell.row, column=cell.column).font = cell.font.copy()
                                _destws.cell(row=cell.row, column=cell.column).fill = cell.fill.copy()
                                _destws.cell(row=cell.row, column=cell.column).border = cell.border.copy()
                                _destws.cell(row=cell.row, column=cell.column).alignment = cell.alignment.copy()
                                _destws.cell(row=cell.row, column=cell.column).number_format = cell.number_format
                    # Kopiraj širine stupaca
                    for col_letter, dim in _srcws.column_dimensions.items():
                        _destws.column_dimensions[col_letter].width = dim.width
                    # Kopiraj merge cells
                    for mc in _srcws.merged_cells.ranges:
                        _destws.merge_cells(str(mc))

                    # ── TOC positions for total tables ──
                    if add_toc:
                        _toc_r = 1
                        for tbl in tables:
                            _gi = tbl.get('_idx')
                            if _gi is not None:
                                toc_positions.setdefault(_gi, []).append({
                                    'sheet': sname, 'row': _toc_r, 'cat': 'total'})
                            n_data = len(tbl['rows'])
                            has_caption = bool(tbl.get('caption', ''))
                            _toc_r += 1 + 1 + n_data + (1 if has_caption else 0) + 1

                _twb.close()
                os.unlink(_tf_path)

            elif out_def['type'] == 'krizanje':
                # ── KRIŽANJE output → banner tablice ──
                banner_vars = out_def.get('banner_vars', [])
                tbl_indices = out_def.get('table_indices', [])
                show_sig = out_def.get('show_sig', True)
                show_sig_total = out_def.get('show_sig_total', False)

                if not banner_vars or not tbl_indices:
                    all_errors.append(f"Output '{out_def['sheet_name']}': nema banner varijabli ili tablica")
                    continue

                col_map = build_column_map(work_df)

                # Sheet bez sig-a (uvijek)
                ws_name = _unique_name(out_def['sheet_name'])
                ws = wb.create_sheet(title=ws_name)

                # Sheet sa sig-om (ako show_sig=True)
                ws_sig = None
                if show_sig:
                    sig_name = _unique_name(out_def['sheet_name'] + '_sig')
                    ws_sig = wb.create_sheet(title=sig_name)

                # Sheet sa sig_total (ako show_sig_total=True)
                ws_sig_total = None
                if show_sig_total:
                    sig_total_name = _unique_name(out_def['sheet_name'] + '_sig_total')
                    ws_sig_total = wb.create_sheet(title=sig_total_name)

                current_row = 1
                current_row_sig = 1
                current_row_st = 1

                # Track kriz output for TOC
                if add_toc:
                    _ki = len(toc_kriz_sheets)
                    toc_kriz_sheets.append({
                        'plain': ws_name,
                        'sig': sig_name if show_sig else None,
                        'sigT': sig_total_name if show_sig_total else None,
                    })

                for ti in tbl_indices:
                    title_line = titles[ti]
                    var_line = variables[ti]  # type: ignore[index]
                    table_type = get_table_type(title_line)
                    table_title = get_table_title(title_line)
                    table_num = ti + start_num

                    # Per-table: compute crosstab for each banner var
                    crosstabs = []
                    for break_var in banner_vars:
                        # Preskoči samo-križanje
                        if table_type == 's' and var_line.strip().lower() == break_var.lower():
                            continue
                        if break_var.lower() in [v.strip().lower() for v in var_line.split()]:
                            continue

                        try:
                            if table_type == 's':
                                xt = make_crosstab_simple(
                                    work_df, var_line.strip(), break_var,
                                    meta, col_map, weight_col)
                            elif table_type in ('k', 'd'):
                                xt = make_crosstab_mr(
                                    work_df, var_line, break_var,
                                    meta, col_map, table_type, weight_col)
                            elif table_type in ('n', 'm'):
                                xt = make_crosstab_numeric(
                                    work_df, var_line, break_var,
                                    meta, col_map, table_type == 'n', weight_col)
                            else:
                                continue
                            crosstabs.append(xt)
                        except Exception as e:
                            all_errors.append(f"Kriz T{table_num} × {break_var}: {e}")

                    if not crosstabs:
                        continue

                    # Merge into banner
                    banner = merge_crosstabs_banner(crosstabs)
                    title_str = table_title
                    total_xt += 1

                    # ── TOC positions for krizanje table ──
                    if add_toc:
                        toc_positions.setdefault(ti, []).append({
                            'sheet': ws_name, 'row': current_row, 'cat': 'kriz', 'kriz_idx': _ki})
                        if ws_sig is not None:
                            toc_positions.setdefault(ti, []).append({
                                'sheet': sig_name, 'row': current_row_sig, 'cat': 'kriz_sig', 'kriz_idx': _ki})
                        if ws_sig_total is not None:
                            toc_positions.setdefault(ti, []).append({
                                'sheet': sig_total_name, 'row': current_row_st, 'cat': 'kriz_sigT', 'kriz_idx': _ki})

                    # Write to plain sheet (no sig)
                    current_row = write_banner_to_sheet(
                        ws, banner, title_str,
                        start_row=current_row, show_sig=False,
                        design=table_design)
                    current_row += 2

                    # Write to sig sheet (with sig)
                    if ws_sig is not None:
                        current_row_sig = write_banner_to_sheet(
                            ws_sig, banner, title_str,
                            start_row=current_row_sig, show_sig=True,
                            design=table_design)
                        current_row_sig += 2

                    # Write to sig_total sheet (sig on Total column)
                    if ws_sig_total is not None:
                        current_row_st = write_banner_to_sheet(
                            ws_sig_total, banner, title_str,
                            start_row=current_row_st, show_sig=True,
                            show_sig_total=True, design=table_design)
                        current_row_st += 2

        # ── TOC sheet (question-centric) ──
        if add_toc and toc_rows:
            import re as _re
            from openpyxl.styles import Font as _Font, PatternFill as _Fill, Alignment as _Align, Border, Side
            from spss_tables import _get_theme
            _toc_t = _get_theme(table_design)
            toc_ws = wb.create_sheet("TOC (BETA)", 0)

            _hfont = _Font(name='Calibri', size=11, bold=True, color=_toc_t.get('toc_header_color', _toc_t['title_color']))
            _hfill = _Fill(start_color=_toc_t['toc_header_fill'], end_color=_toc_t['toc_header_fill'], fill_type='solid')
            _lfont = _Font(name='Calibri', size=10, color=_toc_t['toc_link_color'], underline='single')
            _dfont = _Font(name='Calibri', size=10, color=_toc_t['data_color'])
            _even_fill = _Fill(start_color=_toc_t['toc_even_fill'], end_color=_toc_t['toc_even_fill'], fill_type='solid')
            _brd = Border(bottom=Side(style='thin', color=_toc_t['toc_line_color']))

            def _var_name(title):
                m = _re.match(r'^(\S+)', title.strip())
                return m.group(1).rstrip('.:') if m else title[:20]

            def _make_link(ws, row, col, text, sheet, cell_row, fill=None):
                safe = sheet.replace("'", "''")
                c = ws.cell(row=row, column=col, value=text)
                c.hyperlink = f"#'{safe}'!A{cell_row}"
                c.font = _lfont
                c.border = _brd
                if fill: c.fill = fill
                c.alignment = _Align(horizontal='center')

            # Build dynamic columns: fixed + per-kriz output
            fixed_headers = ['#', 'Pitanje', 'Total', 'MEAN', 'T2B']
            kriz_cols = []  # (header, kriz_idx, cat)
            for ki, ks in enumerate(toc_kriz_sheets):
                kriz_cols.append((ks['plain'], ki, 'kriz'))
                if ks['sig']:
                    kriz_cols.append((ks['sig'], ki, 'kriz_sig'))
                if ks['sigT']:
                    kriz_cols.append((ks['sigT'], ki, 'kriz_sigT'))
            all_headers = fixed_headers + [kc[0] for kc in kriz_cols]

            _hbrd = Border(bottom=Side(style='medium', color='1D1D1B'))

            # Write header row
            for ci, h in enumerate(all_headers, 1):
                c = toc_ws.cell(row=1, column=ci, value=h)
                c.font = _hfont
                c.fill = _hfill
                c.alignment = _Align(horizontal='center', vertical='center')
                c.border = _hbrd

            # Write data rows
            for ri, qrow in enumerate(toc_rows, 2):
                _row_fill = _even_fill if ri % 2 == 0 else None
                # #
                c = toc_ws.cell(row=ri, column=1, value=ri - 1)
                c.font = _dfont; c.border = _brd
                if _row_fill: c.fill = _row_fill
                c.alignment = _Align(horizontal='center')

                # Pitanje
                c = toc_ws.cell(row=ri, column=2, value=qrow['title'])
                c.font = _dfont; c.border = _brd
                if _row_fill: c.fill = _row_fill

                # Total (col 3)
                _ridx = qrow.get('regular_idx')
                if _ridx is not None:
                    _tp = [p for p in toc_positions.get(_ridx, []) if p['cat'] == 'total']
                    if _tp:
                        _make_link(toc_ws, ri, 3, _var_name(qrow['title']),
                                   _tp[0]['sheet'], _tp[0]['row'], _row_fill)
                    else:
                        c = toc_ws.cell(row=ri, column=3); c.border = _brd
                        if _row_fill: c.fill = _row_fill
                else:
                    c = toc_ws.cell(row=ri, column=3); c.border = _brd
                    if _row_fill: c.fill = _row_fill

                # MEAN (col 4)
                _midx = qrow.get('mean_idx')
                if _midx is not None:
                    _tp = [p for p in toc_positions.get(_midx, []) if p['cat'] == 'total']
                    if _tp:
                        _mt = get_table_title(titles[_midx])
                        _make_link(toc_ws, ri, 4, _var_name(_mt),
                                   _tp[0]['sheet'], _tp[0]['row'], _row_fill)
                    else:
                        c = toc_ws.cell(row=ri, column=4); c.border = _brd
                        if _row_fill: c.fill = _row_fill
                else:
                    c = toc_ws.cell(row=ri, column=4); c.border = _brd
                    if _row_fill: c.fill = _row_fill

                # T2B (col 5)
                _tidx = qrow.get('t2b_idx')
                if _tidx is not None:
                    _tp = [p for p in toc_positions.get(_tidx, []) if p['cat'] == 'total']
                    if _tp:
                        _make_link(toc_ws, ri, 5, _var_name(qrow['title']),
                                   _tp[0]['sheet'], _tp[0]['row'], _row_fill)
                    else:
                        c = toc_ws.cell(row=ri, column=5); c.border = _brd
                        if _row_fill: c.fill = _row_fill
                else:
                    c = toc_ws.cell(row=ri, column=5); c.border = _brd
                    if _row_fill: c.fill = _row_fill

                # Krizanje columns
                for kci, (kname, kidx, kcat) in enumerate(kriz_cols):
                    col_num = 6 + kci
                    _found = False
                    for _try_idx in [_ridx, _tidx, _midx]:
                        if _try_idx is None:
                            continue
                        _tp = [p for p in toc_positions.get(_try_idx, [])
                               if p['cat'] == kcat and p.get('kriz_idx') == kidx]
                        if _tp:
                            _make_link(toc_ws, ri, col_num, _var_name(qrow['title']),
                                       _tp[0]['sheet'], _tp[0]['row'], _row_fill)
                            _found = True
                            break
                    if not _found:
                        c = toc_ws.cell(row=ri, column=col_num); c.border = _brd
                        if _row_fill: c.fill = _row_fill

            # Column widths
            toc_ws.column_dimensions['A'].width = 5
            toc_ws.column_dimensions['B'].width = 55
            from openpyxl.utils import get_column_letter as _gcl
            for ci in range(3, len(all_headers) + 1):
                toc_ws.column_dimensions[_gcl(ci)].width = 14
            toc_ws.freeze_panes = 'C2'

        # Spremi workbook
        if not wb.worksheets:
            wb.create_sheet("Sheet1")
        wb.save(tmp_path)
        wb.close()

        elapsed = time.time() - t_start

        with open(tmp_path, 'rb') as f:
            excel_bytes = f.read()
        os.unlink(tmp_path)

        progress.progress(100, text="Gotovo!")

        # Rezultati
        st.divider()

        col_res1, col_res2, col_res3, col_res4 = st.columns(4)
        col_res1.metric("Total tablica", total_tables)
        col_res2.metric("Banner tablica", total_xt)
        col_res3.metric("Vrijeme", f"{elapsed:.2f}s")
        col_res4.metric("Sheetova", len(existing_sheets))

        if all_errors:
            with st.expander(f"⚠️ {len(all_errors)} grešaka", expanded=True):
                for e in all_errors:
                    st.warning(e)

        # ── Download gumb ──
        sav_base = os.path.splitext(sav_file.name)[0]  # type: ignore[union-attr]
        # Strip trailing version suffix (e.g. _v4, _V12) from project name
        sav_base = re.sub(r'_[vV]\d+$', '', sav_base)
        pond_part = '_pond' if use_weight else ''
        output_name = f"{sav_base}_tables{pond_part}_v1.xlsx"

        st.download_button(
            label=f"⬇️ Preuzmi {output_name}",
            data=excel_bytes,
            file_name=output_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True,
        )

        st.success("✅ Tablice uspješno generirane!")


if __name__ == '__main__':
    main()
